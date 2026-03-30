"""
三方对比 — 字段对齐、合规判定、写对比 Excel

将供应商检测值（extract_named_fields）与内部检测值+标准值（extract_internal_fields）
按标准字段名对齐，输出 7 列对比 Excel。
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .constants import MECH_FIELD_PATTERNS
from .internal_extractor import match_to_standard_name
from .judgment import check_value, parse_requirement
from .utils import is_skip_value

_ORDERED_FIELDS = [f for f, _ in MECH_FIELD_PATTERNS]

# 不输出到对比报告的字段（不在质量证明书对比范围内）
EXCLUDE_FROM_COMPARISON: set[str] = {"冷弯"}

_COLS = ["检测项", "标准值", "供应商数值", "内部检测值", "是否合格", "供应商不合格原因", "内部不合格原因"]
_COL_WIDTHS = [14, 16, 14, 14, 10, 22, 22]


def _supplier_reason(actual: str, rule: dict) -> str:
    """生成供应商不合格原因文本。"""
    result = check_value(actual, rule)
    if result is False:
        t = rule["type"]
        if t == "range":
            return f"{actual}<{rule['min']}或>{rule['max']}"
        elif t == "gte":
            return f"{actual}<{rule['value']}"
        elif t == "lte":
            return f"{actual}>{rule['value']}"
        elif t == "eq":
            exp = rule["value"]
            exp_s = str(int(exp)) if float(exp).is_integer() else str(exp)
            return f"{actual}≠{exp_s}"
    return ""


def _normalize_supplier(supplier_fields: list[dict]) -> dict[str, str]:
    """
    将 extract_named_fields 返回的列表转为 {标准字段名: 供应商数值} 字典。
    每个 dict 中除 "方向" 外的键已经是标准字段名（如 "屈服强度"）。
    取第一个有效方向（纵向优先）的值。
    """
    priority = {"纵向": 0, "横向": 1, "切向": 2}
    sorted_recs = sorted(supplier_fields, key=lambda r: priority.get(r.get("方向", ""), 99))

    result: dict[str, str] = {}
    for rec in sorted_recs:
        for k, v in rec.items():
            if k == "方向":
                continue
            std_name = match_to_standard_name(k) or k
            if std_name not in result and v != "—" and not is_skip_value(v):
                result[std_name] = v

    # ── OCR 误识别修正 ────────────────────────────────────────────
    # 当 硬度HBW 未提取到值，但 冷弯 槽位有数值时（系列位置滑移导致），
    # 将该值归入 硬度HBW（质量证明书中冷弯一般不附数值，出现数字必为硬度误归）
    hbw = result.get("硬度HBW", "")
    cold = result.get("冷弯", "")
    if (not hbw or is_skip_value(hbw)) and cold and not is_skip_value(cold):
        try:
            first_num = float(re.split(r'[/\s]', cold)[0])
            if 50 <= first_num <= 700:          # 硬度值合理范围
                result["硬度HBW"] = cold
        except (ValueError, TypeError):
            pass
    result.pop("冷弯", None)                    # 冷弯不参与对比
    return result


def build_comparison(
    supplier_fields: list[dict],
    internal_fields: list[dict],
) -> list[dict]:
    """
    按标准字段名对齐供应商检测值与内部检测+标准值，返回对比行列表。

    每行格式：
    {
        "检测项": str,
        "标准值": str,
        "供应商数值": str,
        "内部检测值": str,
        "是否合格": str,          # "合格" / "不合格" / "待确认"
        "供应商不合格原因": str,
        "内部不合格原因": str,
    }
    """
    supplier_map = _normalize_supplier(supplier_fields)
    internal_map: dict[str, dict] = {item["检测项"]: item for item in internal_fields}

    all_fields: list[str] = []
    seen: set[str] = set()
    for f in _ORDERED_FIELDS:
        if f in EXCLUDE_FROM_COMPARISON:
            continue
        if f in supplier_map or f in internal_map:
            all_fields.append(f)
            seen.add(f)
    for f in list(internal_map.keys()) + list(supplier_map.keys()):
        if f not in seen and f not in EXCLUDE_FROM_COMPARISON:
            all_fields.append(f)
            seen.add(f)

    rows: list[dict] = []
    for field in all_fields:
        supplier_val = supplier_map.get(field, "")
        internal_item = internal_map.get(field, {})
        standard_val = internal_item.get("标准值", "")
        internal_val = internal_item.get("内部检测值", "")
        internal_reason = internal_item.get("内部不合格原因", "")
        internal_pass = internal_item.get("内部合格", None)

        supplier_reason = ""
        supplier_pass: bool | None = None
        if supplier_val and standard_val:
            rule = parse_requirement(standard_val)
            if rule:
                supplier_pass = check_value(supplier_val, rule)
                if supplier_pass is False:
                    supplier_reason = _supplier_reason(supplier_val, rule)
            else:
                supplier_pass = None
        elif not supplier_val:
            supplier_pass = None

        if supplier_pass is None or internal_pass is None:
            verdict = "待确认"
        elif supplier_pass is True and internal_pass is True:
            verdict = "合格"
        else:
            verdict = "不合格"

        rows.append({
            "检测项": field,
            "标准值": standard_val or "—",
            "供应商数值": supplier_val or "—",
            "内部检测值": internal_val or "—",
            "是否合格": verdict,
            "供应商不合格原因": supplier_reason or "—",
            "内部不合格原因": internal_reason or "—",
        })

    return rows


def write_comparison_excel(rows: list[dict], output_path: str):
    """
    将对比行写入 7 列 Excel 文件。
    - 不合格行：红色背景 #FFCCCC
    - 待确认行：黄色背景 #FFF2CC
    - 表头：灰色背景 #D9D9D9 + 加粗
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "三方对比"

    FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9D9D9")
    FILL_FAIL = PatternFill(fill_type="solid", fgColor="FFCCCC")
    FILL_PENDING = PatternFill(fill_type="solid", fgColor="FFF2CC")
    FONT_HEADER = Font(bold=True)
    FONT_FAIL = Font(bold=True, color="CC0000")
    CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for c, (col_name, width) in enumerate(zip(_COLS, _COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    for r, row in enumerate(rows, 2):
        verdict = row.get("是否合格", "")
        for c, col_name in enumerate(_COLS, 1):
            cell = ws.cell(row=r, column=c, value=row.get(col_name, ""))
            cell.alignment = CENTER
            if verdict == "不合格":
                cell.fill = FILL_FAIL
                if col_name in ("供应商不合格原因", "内部不合格原因", "是否合格"):
                    cell.font = FONT_FAIL
            elif verdict == "待确认":
                cell.fill = FILL_PENDING

    wb.save(output_path)
    print(f"  对比报告已保存：{Path(output_path).name}（共 {len(rows)} 条检测项）")
