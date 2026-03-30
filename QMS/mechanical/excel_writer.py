"""
机械性能提取 — HTML 表格写入 Excel
"""

from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import clean_cell_text


def html_table_to_sheet(
    ws, html: str, row_offset: int = 1,
) -> tuple[int, dict[tuple[int, int], int]]:
    """
    将 HTML <table> 写入 openpyxl worksheet，处理 colspan/rowspan。
    返回 (最后行号, td_coord_map)。
    """
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return row_offset, {}
    occupied: dict[tuple[int, int], bool] = {}
    td_coord_map: dict[tuple[int, int], int] = {}
    td_index = 0
    excel_row = row_offset
    for tr in table.find_all("tr"):
        excel_col = 1
        for td in tr.find_all(["td", "th"]):
            while occupied.get((excel_row, excel_col)):
                excel_col += 1
            colspan = int(td.get("colspan", 1))
            rowspan = int(td.get("rowspan", 1))
            text = clean_cell_text(td.get_text(separator=" ", strip=True))
            ws.cell(row=excel_row, column=excel_col, value=text)
            end_row = excel_row + rowspan - 1
            end_col = excel_col + colspan - 1
            for r in range(excel_row, end_row + 1):
                for c in range(excel_col, end_col + 1):
                    td_coord_map[(r, c)] = td_index
            td_index += 1
            if colspan > 1 or rowspan > 1:
                ws.merge_cells(
                    start_row=excel_row, start_column=excel_col,
                    end_row=end_row, end_column=end_col,
                )
                for r in range(excel_row, end_row + 1):
                    for c in range(excel_col, end_col + 1):
                        if r != excel_row or c != excel_col:
                            occupied[(r, c)] = True
            cell = ws.cell(row=excel_row, column=excel_col)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            excel_col += colspan
        excel_row += 1
    return excel_row, td_coord_map


def apply_header_style(ws, max_row: int, max_col: int):
    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill


def auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len * 1.2, 8), 50)
