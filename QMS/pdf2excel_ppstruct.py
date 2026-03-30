"""
PDF 表格识别 → Excel 输出（PP-Structure V3 版）

与 pdf2excel.py（逐行文字块方案）的对比版本。
使用 PPStructureV3 直接识别表格结构，输出 HTML 后转 Excel，
单元格内多行文字自动合并，colspan/rowspan 自动还原。

用法：
    conda activate paddleocr310
    python pdf2excel_ppstruct.py -i <PDF路径> [-o <输出目录>]

示例：
    python pdf2excel_ppstruct.py -i "/Users/project/QMS/PaddleOCR/文件/机械性能复检报告.pdf"

输出文件：
    <pdf名>_ppstruct_table.xlsx  — 还原表格结构的 Excel（含单元格合并）
    <pdf名>_ppstruct_page_N.html — 每页表格的原始 HTML（供调试）
"""

import argparse
import os
import re
import shutil
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

os.environ["PADDLE_PDX_MODEL_SOURCE"] = "modelscope"
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

import paddle
paddle.set_device("cpu")

import pypdfium2 as pdfium
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
import numpy as np
from paddleocr import PPStructureV3
from PIL import Image, ImageDraw

from ocr_config import load_config
from ocr_db import build_record, ensure_nonconformity_table, insert_nonconformity_records

MAX_IMAGE_WIDTH = 1600


# ─── 命令行参数 ───────────────────────────────────────────────────────────────

def parse_args():
    cfg = load_config()
    parser = argparse.ArgumentParser(
        description="PDF / 图片表格识别（PP-Structure V3 版）",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "-i", "--input", required=True,
        help="输入文件路径，支持 PDF / JPG / JPEG / PNG",
    )
    parser.add_argument(
        "-o", "--output",
        default=cfg["output"]["output_dir"],
        help="输出目录",
    )
    return parser.parse_args()


# ─── 公章去除 ─────────────────────────────────────────────────────────────────

def remove_red_stamp(img: Image.Image) -> Image.Image:
    """
    将图片中的红色区域（公章印章）替换为白色，减少 OCR 干扰。

    判断条件（RGB）：
      R > 150  且  G < 90  且  B < 90
    适用于常见的红色圆形公章；蓝色/黑色章不受影响。
    """
    arr = np.array(img.convert("RGB"), dtype=np.uint8)
    r, g, b = arr[:, :, 0], arr[:, :, 1], arr[:, :, 2]
    mask = (r > 150) & (g < 90) & (b < 90)
    arr[mask] = [255, 255, 255]
    result = Image.fromarray(arr)
    removed = int(mask.sum())
    if removed > 0:
        print(f"  [去章] 已清除红色区域 {removed} 个像素")
    return result


# ─── PDF 渲染 ─────────────────────────────────────────────────────────────────

def pdf_to_images(
    pdf_path: str,
    tmp_dir: str | Path,
    remove_stamp: bool = True,
    max_width: int = MAX_IMAGE_WIDTH,
) -> list[tuple[int, str]]:
    """
    将 PDF 每页渲染为临时图片，返回 [(页码, 图片路径), ...]。
    图片保存到 tmp_dir（临时目录），识别完成后由调用方统一清理。
    remove_stamp=True 时自动去除红色公章。
    """
    pdf      = pdfium.PdfDocument(pdf_path)
    tmp      = Path(tmp_dir)
    tmp.mkdir(parents=True, exist_ok=True)
    pages    = []
    for i in range(len(pdf)):
        page  = pdf[i]
        w, h  = page.get_size()
        scale = min(max_width / w, 2.0)
        img   = page.render(scale=scale).to_pil()
        if remove_stamp:
            img = remove_red_stamp(img)
        p     = tmp / f"_page_{i+1}.png"
        img.save(str(p))
        print(f"  第 {i+1} 页 → {img.size[0]}×{img.size[1]}  准备完成")
        pages.append((i + 1, str(p)))
    return pages


def image_to_pages(
    img_path: str,
    tmp_dir: str | Path,
    remove_stamp: bool = True,
) -> list[tuple[int, str]]:
    """
    将单张图片（JPG/JPEG/PNG）预处理后保存到 tmp_dir（临时目录），
    返回 [(1, 图片路径)]。识别完成后由调用方统一清理。
    remove_stamp=True 时自动去除红色公章。
    """
    tmp  = Path(tmp_dir)
    tmp.mkdir(parents=True, exist_ok=True)
    dest = tmp / "_page_1.png"
    img  = Image.open(img_path).convert("RGB")
    w, h = img.size
    if w > MAX_IMAGE_WIDTH:
        scale = MAX_IMAGE_WIDTH / w
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    if remove_stamp:
        img = remove_red_stamp(img)
    img.save(str(dest))
    print(f"  图片输入 → {img.size[0]}×{img.size[1]}  准备完成")
    return [(1, str(dest))]


# ─── PP-Structure V3 识别 ─────────────────────────────────────────────────────

def _parse_cell_bboxes(s: str) -> list[list[int]] | None:
    """
    从 PPStructureV3 table_res_list 对象字符串中提取单元格精确坐标列表。
    支持两种常见格式：
      cell_bbox: [[x1,y1,x2,y2], ...]
      cell_bboxes: [[x1,y1,x2,y2], ...]
    返回 [[x1,y1,x2,y2], ...] 或 None。
    """
    m = re.search(r"cell_bbox[s]?[:\s\t]+(\[\s*\[.+?\]\s*\])", s, re.DOTALL)
    if not m:
        return None
    try:
        raw = m.group(1)
        rows = re.findall(r"\[([^\[\]]+)\]", raw)
        result = []
        for row in rows:
            vals = re.findall(r"-?\d+\.?\d*", row)
            if len(vals) >= 4:
                result.append([int(float(v)) for v in vals[:4]])
        return result if result else None
    except Exception:
        return None


def _seq_for_ocr(x) -> list:
    """
    将 Paddle 返回的 rec_texts / rec_boxes 等转为 Python 列表。
    禁止对 ndarray 做 `x or []`、`if x:` 等布尔判断，否则会触发 ambiguous truth value。
    """
    if x is None:
        return []
    if isinstance(x, np.ndarray):
        return x.tolist()
    if isinstance(x, (list, tuple)):
        return list(x)
    return [x]


def _rec_box_to_xyxy(box) -> list[int] | None:
    """将 rec_boxes / rec_polys 单项转为轴对齐 [x1,y1,x2,y2]。"""
    if box is None:
        return None
    arr = np.asarray(box, dtype=float).reshape(-1)
    if arr.size < 4:
        return None
    if arr.size == 4:
        x1, y1, x2, y2 = (float(arr[i]) for i in range(4))
    else:
        pts = arr.reshape(-1, 2)
        x1, x2 = float(pts[:, 0].min()), float(pts[:, 0].max())
        y1, y2 = float(pts[:, 1].min()), float(pts[:, 1].max())
    return [int(round(x1)), int(round(y1)), int(round(x2)), int(round(y2))]


def _extract_ocr_lines_from_pp_result(r: dict) -> list[dict]:
    """
    从单页 PPStructure 结果中提取通用 OCR 文本行及其在原图上的框。
    返回 [{"text": str, "bbox": [x1,y1,x2,y2]}, ...]，与表格识别同源，用于紧贴数值画标注框。
    """
    ocr = r.get("overall_ocr_res")
    if not isinstance(ocr, dict):
        return []
    texts = _seq_for_ocr(ocr.get("rec_texts"))
    boxes = _seq_for_ocr(ocr.get("rec_boxes"))
    polys = _seq_for_ocr(ocr.get("rec_polys"))
    lines: list[dict] = []
    for i, t in enumerate(texts):
        if t is None:
            continue
        bb = None
        if i < len(boxes):
            bb = _rec_box_to_xyxy(boxes[i])
        if bb is None and i < len(polys):
            bb = _rec_box_to_xyxy(polys[i])
        if bb is None:
            continue
        lines.append({"text": str(t), "bbox": bb})
    if lines:
        print(f"    → OCR 文本行：{len(lines)} 条（不合格标注优先贴合识别框）")
    return lines


def _filter_ocr_lines_to_table(
    ocr_lines: list[dict],
    table_bbox: list | None,
    margin: int = 8,
) -> list[dict]:
    """
    只保留几何中心落在 table_bbox（可扩边）内的 OCR 行，与「该表」识别坐标一致。
    若过滤后为空（外框偏差大），退回全页 OCR，避免无匹配。
    """
    if not ocr_lines:
        return []
    tb = _seq_for_ocr(table_bbox) if table_bbox is not None else []
    if len(tb) < 4:
        return list(ocr_lines)
    tx1, ty1, tx2, ty2 = (float(tb[i]) for i in range(4))
    out: list[dict] = []
    for L in ocr_lines:
        bb = L.get("bbox")
        if not bb or len(bb) < 4:
            continue
        cx = (bb[0] + bb[2]) / 2
        cy = (bb[1] + bb[3]) / 2
        if (tx1 - margin) <= cx <= (tx2 + margin) and (ty1 - margin) <= cy <= (ty2 + margin):
            out.append(L)
    return out if out else list(ocr_lines)


def extract_tables_from_page(engine: PPStructureV3, img_path: str) -> tuple[list[dict], list[dict]]:
    """
    对单页图片运行 PPStructureV3，返回 (表格列表, OCR 文本行列表)。

    表格项：{
      "html": str,
      "table_bbox": [x1, y1, x2, y2] | None,
      "cell_bboxes": [[x1,y1,x2,y2], ...] | None,
      "table_ocr_lines": 表区域内的 OCR 行列表（text+bbox，与 predict 底图同源）,
    }
    OCR 行项：{"text": str, "bbox": [x1,y1,x2,y2]}
    """
    results = list(engine.predict(img_path))
    if not results:
        return [], []

    r = results[0]

    # 与 predict 同一张底图上的全文 OCR（坐标系一致，后续按表裁剪、按格绑定）
    ocr_lines = _extract_ocr_lines_from_pp_result(r)

    # 尝试从 table_res_list 提取每个表格的精确单元格坐标
    tbl_cell_bboxes: list[list[list[int]] | None] = []
    for tbl_res in r.get("table_res_list", []):
        tbl_str = str(tbl_res)
        cell_bboxes = _parse_cell_bboxes(tbl_str)
        # PPStructure V3 的 SingleTableRecognitionResult 是字典风格对象
        # cell_box_list 通过 [] 访问（getattr 无效）
        if cell_bboxes is None:
            try:
                raw_boxes = tbl_res["cell_box_list"]
                if raw_boxes:
                    bboxes = []
                    for b in raw_boxes:
                        arr = np.asarray(b, dtype=float).reshape(-1)
                        if arr.size >= 4:
                            bboxes.append([int(round(float(arr[0]))),
                                           int(round(float(arr[1]))),
                                           int(round(float(arr[2]))),
                                           int(round(float(arr[3])))])
                    if bboxes:
                        cell_bboxes = bboxes
            except Exception:
                pass
        tbl_cell_bboxes.append(cell_bboxes)

    tables = []
    tbl_idx = 0
    for item in r.get("parsing_res_list", []):
        item_str = str(item)
        if "label:\ttable" in item_str or "label: table" in item_str:
            html = ""
            m = re.search(r"content:\t(.*)", item_str, re.DOTALL)
            if m:
                html = m.group(1).strip()
            if not html:
                continue

            # 整个表格的外框 bbox
            table_bbox = None
            bm = re.search(r"bbox[:\s\t]+\[([^\]]+)\]", item_str)
            if bm:
                try:
                    vals = re.findall(r"-?\d+\.?\d*", bm.group(1))
                    if len(vals) >= 4:
                        table_bbox = [int(float(v)) for v in vals[:4]]
                except (ValueError, IndexError):
                    pass

            # 单元格精确坐标
            cell_bboxes = tbl_cell_bboxes[tbl_idx] if tbl_idx < len(tbl_cell_bboxes) else None
            if isinstance(cell_bboxes, np.ndarray):
                cell_bboxes = cell_bboxes.tolist()
            if cell_bboxes is not None and len(cell_bboxes) > 0:
                print(f"    → 表格 cell_bboxes：{len(cell_bboxes)} 个单元格坐标（精确模式）")
            elif table_bbox is not None and len(_seq_for_ocr(table_bbox)) >= 4:
                print(f"    → 表格 bbox：{table_bbox}（均匀估算模式）")
            else:
                print(f"    → 表格坐标：未提取到（图片标注将跳过）")

            scoped = _filter_ocr_lines_to_table(ocr_lines, table_bbox)
            tables.append(
                {
                    "html": html,
                    "table_bbox": table_bbox,
                    "cell_bboxes": cell_bboxes,
                    "table_ocr_lines": scoped,
                }
            )
            tbl_idx += 1
    return tables, ocr_lines


# ─── HTML 表格 → openpyxl Sheet ───────────────────────────────────────────────

# ─── 判定规则引擎 ─────────────────────────────────────────────────────────────

def _normalize_requirement_string(req: str) -> str:
    """
    统一 OCR 常见写法，避免「-50」类要求值解析失败导致整列不参与判定。
    - 全角/长横负号 → ASCII '-'
    - 行首「- 50」→「-50」
    - 末尾温度单位（℃、°C）去掉后再走正则
    """
    s = req.strip()
    for a, b in (
        ("\u2212", "-"),
        ("\u2013", "-"),
        ("\u2014", "-"),
        ("－", "-"),
    ):
        s = s.replace(a, b)
    s = re.sub(r"^-\s+", "-", s)
    s = re.sub(r"^\+\s+", "+", s)
    s = re.sub(r"\s*(?:℃|°\s*C|°C)\s*$", "", s, flags=re.I)
    return s.strip()


def _normalize_actual_number_token(p: str) -> str:
    """实测值片段：负号与单位规范化，保证 float 可解析、eq 可比。"""
    s = p.strip()
    for a, b in (
        ("\u2212", "-"),
        ("\u2013", "-"),
        ("\u2014", "-"),
        ("－", "-"),
    ):
        s = s.replace(a, b)
    s = re.sub(r"^-\s+", "-", s)
    s = re.sub(r"^\+\s+", "+", s)
    s = re.sub(r"\s*(?:℃|°\s*C|°C)\s*$", "", s, flags=re.I)
    return s.strip()


def parse_requirement(req: str) -> dict | None:
    """
    解析要求值字符串，返回判定规则字典。

    支持四种形式：
      区间      686-800   → {"type": "range", "min": 686, "max": 800}
      大于等于  ≥15       → {"type": "gte",   "value": 15}
      小于等于  ≤X        → {"type": "lte",   "value": X}
      精确等于  -50       → {"type": "eq",    "value": -50}（单独负数、无 ≥≤ 前缀）
    """
    req = _normalize_requirement_string(req)
    if not req:
        return None
    # 区间：两个正数用 - 分隔（负数不会匹配，因为负数开头没有纯数字）
    m = re.match(r'^(\d+\.?\d*)\s*[-–]\s*(\d+\.?\d*)$', req)
    if m:
        return {"type": "range", "min": float(m.group(1)), "max": float(m.group(2))}
    # 大于等于：≥ 或 >=
    m = re.match(r'^[≥>]=?\s*(-?\d+\.?\d*)$', req)
    if m:
        return {"type": "gte", "value": float(m.group(1))}
    # 小于等于：≤ 或 <=
    m = re.match(r'^[≤<]=?\s*(-?\d+\.?\d*)$', req)
    if m:
        return {"type": "lte", "value": float(m.group(1))}
    # 精确等于：纯数字或负数（冲击试验温度等「必须为 -50」走此分支，非 -50 判不合格）
    m = re.match(r'^[-+]?\d+\.?\d*$', req)
    if m:
        return {"type": "eq", "value": float(req)}
    return None


def check_value(actual: str, rule: dict) -> bool | None:
    """
    判断实测值是否满足要求规则。
    返回 True=合格 / False=不合格 / None=无法解析（需人工复核）

    支持多值：冲击吸收能量如 "120/124 /124"，以 / 或空格分隔，
    所有子值均需满足要求才判定为合格。
    """
    actual = actual.strip()
    if not actual:
        return None

    parts = re.split(r'[/\s]+', actual)
    values: list[float] = []
    for p in parts:
        p = _normalize_actual_number_token(p)
        if not p:
            continue
        try:
            values.append(float(p))
        except ValueError:
            return None

    if not values:
        return None

    def _ok(v: float) -> bool:
        t = rule["type"]
        if t == "range":
            return rule["min"] <= v <= rule["max"]
        if t == "gte":
            return v >= rule["value"]
        if t == "lte":
            return v <= rule["value"]
        if t == "eq":
            return abs(v - rule["value"]) < 0.01
        return True

    return all(_ok(v) for v in values)


def _rows_data_index(rows_data: list[tuple[int, list[str]]]) -> dict[int, list[str]]:
    return {r: vals for r, vals in rows_data}


def infer_test_item_for_column(
    rows_by_row: dict[int, list[str]],
    req_excel_row: int,
    col_idx: int,
) -> str:
    """
    根据「要求值」行向上追溯同列表头，拼出检测项目名称（如 屈服强度 Yield Strength）。
    """
    skip_exact = {
        "要求值",
        "Standard",
        "standard",
        "单位",
        "Unit",
        "Unit Symbol",
        "实测值",
        "Actual",
        "试样编号",
        "Sample No.",
        "Sample No",
    }
    parts: list[str] = []
    r = req_excel_row - 1
    steps = 0
    while r >= 1 and steps < 12:
        vals = rows_by_row.get(r)
        if vals is None or col_idx >= len(vals):
            r -= 1
            steps += 1
            continue
        cell = vals[col_idx].strip()
        if not cell or cell in skip_exact:
            r -= 1
            steps += 1
            continue
        parts.append(cell)
        if len(parts) >= 8:
            break
        r -= 1
        steps += 1
    if not parts:
        r = req_excel_row - 1
        while r >= 1:
            vals = rows_by_row.get(r)
            if vals and col_idx < len(vals):
                c = vals[col_idx].strip()
                if c and c not in skip_exact:
                    return c[:768]
            r -= 1
        return f"列{get_column_letter(col_idx + 1)}"
    return " ".join(reversed(parts))[:768]


def apply_value_judgment(ws) -> list[dict]:
    """
    遍历已写入的 worksheet，自动找到「要求值」行和「试样」行，
    按列对比实测值与要求值，并对单元格着色：
      - 不合格：红色背景 #FFCCCC + 红色加粗字体
      - 无法解析：黄色背景 #FFF2CC（需人工复核）

    返回：不合格记录列表，每项为 dict：
      excel_row, excel_col, actual, reason, rule_type, standard_text, sample_no, test_item
    """
    FILL_FAIL    = PatternFill(fill_type="solid", fgColor="FFCCCC")
    FILL_UNKNOWN = PatternFill(fill_type="solid", fgColor="FFF2CC")
    FONT_FAIL    = Font(bold=True, color="CC0000")

    max_row = ws.max_row
    max_col = ws.max_column

    # 构建行列文本矩阵（合并单元格非主格值为空串）
    rows_data: list[tuple[int, list[str]]] = []
    for r in range(1, max_row + 1):
        vals = []
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            vals.append(str(raw).strip() if raw is not None else "")
        rows_data.append((r, vals))

    # 定位「要求值」行
    req_excel_row: int | None = None
    req_vals: list[str] = []
    for excel_row, row_vals in rows_data:
        if any("要求值" in v or "Standard" in v for v in row_vals):
            req_excel_row = excel_row
            req_vals = row_vals
            break

    if req_excel_row is None:
        print("  [判定] 未找到「要求值」行，跳过合规判定。")
        return []

    # 预解析每列的要求规则（列索引 0-based）
    col_rules: dict[int, dict] = {}
    for col_idx, req in enumerate(req_vals):
        rule = parse_requirement(req)
        if rule:
            col_rules[col_idx] = rule

    if not col_rules:
        print("  [判定] 要求值行中未解析到有效规则，跳过。")
        return []

    rows_by_row = _rows_data_index(rows_data)

    # 定位试样行（第一个非空值符合 "数字-数字" 格式，如 22-9637）
    judged = 0
    failures: list[dict] = []
    for excel_row, row_vals in rows_data:
        if excel_row <= req_excel_row:
            continue
        first_val = next((v for v in row_vals if v), "")
        if not re.match(r'^\d{2,}[-–]\d{4,}', first_val):
            continue

        # 对有规则的列逐一判定
        for col_idx, rule in col_rules.items():
            actual = row_vals[col_idx] if col_idx < len(row_vals) else ""
            if not actual:
                continue

            result = check_value(actual, rule)
            cell = ws.cell(row=excel_row, column=col_idx + 1)

            if result is False:
                cell.fill = FILL_FAIL
                cell.font = FONT_FAIL
                t = rule["type"]
                if t == "range":
                    reason = f"{actual}<{rule['min']}或>{rule['max']}"
                elif t == "gte":
                    reason = f"{actual}<{rule['value']}"
                elif t == "lte":
                    reason = f"{actual}>{rule['value']}"
                elif t == "eq":
                    exp = rule["value"]
                    exp_s = str(int(exp)) if float(exp).is_integer() else str(exp)
                    reason = f"{actual}≠{exp_s}"
                else:
                    reason = f"{actual}≠{rule['value']}"
                standard_text = (
                    req_vals[col_idx] if col_idx < len(req_vals) else ""
                )
                test_item = infer_test_item_for_column(
                    rows_by_row, req_excel_row, col_idx
                )
                failures.append(
                    {
                        "excel_row": excel_row,
                        "excel_col": col_idx + 1,
                        "actual": actual,
                        "reason": reason,
                        "rule_type": t,
                        "standard_text": standard_text,
                        "sample_no": first_val,
                        "test_item": test_item,
                    }
                )
                judged += 1
            elif result is None:
                cell.fill = FILL_UNKNOWN

    print(f"  [判定] 完成，不合格单元格：{judged} 个（红色标注）。")
    return failures


def _col_idx(col: int) -> str:
    return get_column_letter(col)


def clean_cell_text(text: str) -> str:
    """
    清理单元格文字中的多余空格。

    - 数字型内容（仅含数字、小数点、负号、斜杠、百分号）：去除所有内部空格
      例：'7 26' → '726'，'1 20/124 /124' → '120/124/124'
    - 文字型内容：将连续多个空格合并为一个
    """
    text = text.strip()
    if re.match(r'^[\d\s./\-\+%]+$', text):
        return re.sub(r'\s+', '', text)
    return re.sub(r'\s{2,}', ' ', text)


def html_table_to_sheet(
    ws, html: str, row_offset: int = 1
) -> tuple[int, dict[tuple[int, int], int]]:
    """
    将 HTML <table> 写入 openpyxl worksheet。
    - 处理 colspan / rowspan 单元格合并
    - 返回 (最后行号, td_coord_map)
      td_coord_map: {(excel_row, excel_col): td_index}
      td_index 与 cell_bboxes 列表顺序一致，用于精确图片标注。

    row_offset: 从第几行开始写（用于多页拼接时追加）
    """
    soup  = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return row_offset, {}

    occupied: dict[tuple[int, int], bool] = {}
    td_coord_map: dict[tuple[int, int], int] = {}
    td_index = 0

    excel_row = row_offset
    for tr in table.find_all("tr"):
        excel_col = 1
        for td in tr.find_all(["td", "th"]):
            while occupied.get((excel_row, excel_col)):
                excel_col += 1

            colspan = int(td.get("colspan", 1))
            rowspan = int(td.get("rowspan", 1))
            text    = clean_cell_text(td.get_text(separator=" ", strip=True))

            ws.cell(row=excel_row, column=excel_col, value=text)

            end_row = excel_row + rowspan - 1
            end_col = excel_col + colspan - 1

            # 合并区域内每个 (行,列) 都映射到同一 td_index，便于与 openpyxl 读格一致
            for r in range(excel_row, end_row + 1):
                for c in range(excel_col, end_col + 1):
                    td_coord_map[(r, c)] = td_index
            td_index += 1

            if colspan > 1 or rowspan > 1:
                ws.merge_cells(
                    start_row=excel_row, start_column=excel_col,
                    end_row=end_row,     end_column=end_col,
                )
                for r in range(excel_row, end_row + 1):
                    for c in range(excel_col, end_col + 1):
                        if r != excel_row or c != excel_col:
                            occupied[(r, c)] = True

            cell = ws.cell(row=excel_row, column=excel_col)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

            excel_col += colspan

        excel_row += 1

    return excel_row, td_coord_map


def apply_header_style(ws, max_row: int, max_col: int):
    """对第一行加粗、浅灰背景。"""
    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill


def auto_column_width(ws):
    """自动调整列宽（按最长内容估算）。"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len * 1.2, 8), 50)


# ─── 图片标注 ──────────────────────────────────────────────────────────────────

def _count_html_td_cells(html: str) -> int:
    """统计 HTML 表格中 td/th 总数（与 html_table_to_sheet 遍历顺序一致）。"""
    if not html:
        return 0
    soup  = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return 0
    return sum(len(tr.find_all(["td", "th"])) for tr in table.find_all("tr"))


def _align_cell_bbox_index(td_idx: int, cell_bboxes: list, n_td: int) -> int | None:
    """
    PPStructure 返回的 cell_bboxes 数量有时比 HTML 中 td 多 1（例如多一个表头/角格框），
    会导致 td 索引与框一一对应时整体向左偏一格。此处按数量差自动平移索引。
    """
    n_b = len(cell_bboxes)
    if n_td <= 0 or td_idx < 0:
        return None
    delta = n_b - n_td
    adj   = td_idx
    if delta == 1:
        adj = td_idx + 1
    elif delta == -1 and td_idx > 0:
        adj = td_idx - 1
    if 0 <= adj < n_b:
        return adj
    return None


def _table_grid_metrics(tbl_info: dict) -> tuple[float, float, float, float, float, float, int, int] | None:
    """table 外框 + HTML 逻辑行列数 → 均匀网格参数。返回 (tx1,ty1,tx2,ty2,cell_w,cell_h,n_rows,n_cols)。"""
    table_bbox = tbl_info.get("table_bbox")
    html       = tbl_info.get("html", "")
    tb_list = _seq_for_ocr(table_bbox) if table_bbox is not None else []
    if len(tb_list) < 4 or not html:
        return None
    soup  = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return None
    rows   = table.find_all("tr")
    n_rows = len(rows)
    n_cols = max(
        (sum(int(td.get("colspan", 1)) for td in tr.find_all(["td", "th"])) for tr in rows),
        default=0,
    )
    if n_rows == 0 or n_cols == 0:
        return None
    tx1, ty1, tx2, ty2 = (float(tb_list[i]) for i in range(4))
    cell_w = (tx2 - tx1) / n_cols
    cell_h = (ty2 - ty1) / n_rows
    return tx1, ty1, tx2, ty2, cell_w, cell_h, n_rows, n_cols


def _grid_rect_pixels(
    tbl_info: dict,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[int] | None:
    """Excel 1-based 闭区间 [min_row,max_row]×[min_col,max_col] → 底图像素矩形。"""
    m = _table_grid_metrics(tbl_info)
    if not m:
        return None
    tx1, ty1, _tx2, _ty2, cell_w, cell_h, n_rows, n_cols = m
    r1 = max(0, min(min_row - 1, n_rows - 1))
    r2 = max(0, min(max_row - 1, n_rows - 1))
    c1 = max(0, min(min_col - 1, n_cols - 1))
    c2 = max(0, min(max_col - 1, n_cols - 1))
    if r2 < r1:
        r1, r2 = r2, r1
    if c2 < c1:
        c1, c2 = c2, c1
    x1 = int(tx1 + c1 * cell_w)
    x2 = int(tx1 + (c2 + 1) * cell_w)
    y1 = int(ty1 + r1 * cell_h)
    y2 = int(ty1 + (r2 + 1) * cell_h)
    return [x1, y1, x2, y2]


def _coarse_bbox_from_ws_cell(ws, tbl_info: dict, excel_row: int, excel_col: int) -> list[int] | None:
    """
    结合工作表合并区域：若 (row,col) 在合并块内，粗框覆盖整块（像素与 HTML 表行对齐）。
    """
    min_r = max_r = excel_row
    min_c = max_c = excel_col
    for rng in ws.merged_cells.ranges:
        c1, r1, c2, r2 = range_boundaries(str(rng))
        if r1 <= excel_row <= r2 and c1 <= excel_col <= c2:
            min_r, max_r, min_c, max_c = r1, r2, c1, c2
            break
    return _grid_rect_pixels(tbl_info, min_r, max_r, min_c, max_c)


def _grid_cell_bbox_only(tbl_info: dict, excel_row: int, excel_col: int) -> list[int] | None:
    """
    仅用 table_bbox 与 HTML 表格行列数做均匀网格，得到单元格在底图上的矩形。
    不读取 cell_bboxes，与 PPStructure 输出的 table 外框、识别坐标系一致，供绑定 OCR 行时作粗定位。
    """
    return _grid_rect_pixels(tbl_info, excel_row, excel_row, excel_col, excel_col)


def _get_cell_img_bbox(
    tbl_info: dict,
    td_coord_map: dict,
    excel_row: int,
    excel_col: int,
) -> list[int] | None:
    """
    获取单元格在图片上的精确坐标。
    优先使用 cell_bboxes（精确），不可用时退回均匀网格估算。
    返回 [ix1, iy1, ix2, iy2] 或 None。
    """
    cell_bboxes = tbl_info.get("cell_bboxes")
    td_idx      = td_coord_map.get((excel_row, excel_col))
    html        = tbl_info.get("html", "")

    # 精确模式：用 PPStructureV3 提供的单元格坐标（cell_bboxes 可能是 ndarray，勿用 if cell_bboxes）
    if (
        cell_bboxes is not None
        and len(cell_bboxes) > 0
        and td_idx is not None
    ):
        n_td = _count_html_td_cells(html)
        if n_td > 0:
            adj = _align_cell_bbox_index(td_idx, cell_bboxes, n_td)
        else:
            adj = td_idx if 0 <= td_idx < len(cell_bboxes) else None
        if adj is not None:
            bb = cell_bboxes[adj]
            arr = np.asarray(bb, dtype=float).reshape(-1)
            if arr.size >= 4:
                return [int(round(arr[i])) for i in range(4)]

    # 退回模式：均匀网格估算（与 _grid_cell_bbox_only 相同）
    return _grid_cell_bbox_only(tbl_info, excel_row, excel_col)


def _bbox_intersection_area(a: list[int], b: list[int]) -> int:
    dx = min(a[2], b[2]) - max(a[0], b[0])
    dy = min(a[3], b[3]) - max(a[1], b[1])
    return max(0, dx) * max(0, dy)


def _expand_bbox(bb: list[int], px: int, img_w: int, img_h: int) -> list[int]:
    return [
        max(0, bb[0] - px),
        max(0, bb[1] - px),
        min(img_w, bb[2] + px),
        min(img_h, bb[3] + px),
    ]


def _resolve_ocr_bbox_for_cell(
    table_ocr_lines: list[dict],
    cell_text: str,
    coarse_bbox: list[int] | None,
    img_w: int,
    img_h: int,
    expand_px: int = 28,
) -> list[int] | None:
    """
    在「识别阶段」已截好的表内 OCR 行中，为单元格文本选一条最匹配的 bbox（与底图同源）。
    coarse_bbox 为网格粗框，用于多处同文时的消歧。
    """
    if not table_ocr_lines:
        return None
    ct = clean_cell_text(str(cell_text).strip()) if cell_text else ""
    if not ct:
        return None
    matches = _ocr_lines_for_actual(ct, table_ocr_lines)
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]["bbox"]
    merged = _try_merge_split_matches(ct, matches)
    if merged is not None:
        return merged
    if coarse_bbox is None:
        return matches[0]["bbox"]
    ex = _expand_bbox(coarse_bbox, expand_px, img_w, img_h)
    ch = float(coarse_bbox[3] - coarse_bbox[1])

    def score(L: dict) -> tuple:
        bb = L["bbox"]
        ia = _bbox_intersection_area(bb, ex)
        cy, cx = (bb[1] + bb[3]) / 2, (bb[0] + bb[2]) / 2
        ey = (ex[1] + ex[3]) / 2
        ex_c = (ex[0] + ex[2]) / 2
        bh = float(bb[3] - bb[1])
        area = float((bb[2] - bb[0]) * (bb[3] - bb[1]))
        tall_pen = (1e6 + bh) if (ch > 6 and bh > ch * 2.15) else 0.0
        return (ia, -tall_pen, -area, -abs(cy - ey), -abs(cx - ex_c))

    return max(matches, key=score)["bbox"]


def build_cell_recognition_boxes(ws, tbl_info: dict, td_coord_map: dict, img_size: tuple[int, int]) -> None:
    """
    写入 Excel 后：按 td 合并块，用表内 OCR + 合并感知网格粗框为每个 (行,列) 绑定识别框。
    结果写入 tbl_info["cell_recognition_boxes"]（作兜底，错误标注优先按 actual 现场匹配）。
    """
    img_w, img_h = img_size
    table_ocr = tbl_info.get("table_ocr_lines")
    if not table_ocr:
        table_ocr = []
    by_td: dict[int, list[tuple[int, int]]] = defaultdict(list)
    for (r, c), ti in td_coord_map.items():
        by_td[ti].append((r, c))
    out: dict[tuple[int, int], list[int] | None] = {}
    for _ti, positions in by_td.items():
        r0, c0 = min(positions)
        raw = ws.cell(row=r0, column=c0).value
        text = clean_cell_text(str(raw)) if raw is not None else ""
        coarse = _coarse_bbox_from_ws_cell(ws, tbl_info, r0, c0)
        bb = _resolve_ocr_bbox_for_cell(table_ocr, text, coarse, img_w, img_h)
        for pos in positions:
            out[pos] = bb
    tbl_info["cell_recognition_boxes"] = out


def _ocr_text_matches_actual(actual: str, ocr_text: str) -> bool:
    a = clean_cell_text(actual)
    o = clean_cell_text(ocr_text)
    if not a:
        return False
    if a == o:
        return True
    def squeeze(s: str) -> str:
        s = s.replace("．", ".").replace("，", ",")
        return re.sub(r"\s+", "", s)

    return squeeze(a) == squeeze(o)


def _ocr_lines_for_actual(actual: str, ocr_lines: list[dict]) -> list[dict]:
    """先精确匹配，再无精确时用去空格后的子串匹配（适合 OCR 多字场景）。"""
    if not actual or not ocr_lines:
        return []
    m = [L for L in ocr_lines if _ocr_text_matches_actual(actual, L["text"])]
    if m:
        return m
    sa = clean_cell_text(str(actual).strip())
    if len(sa) < 2:
        return []

    def sq(s: str) -> str:
        return re.sub(r"\s+", "", s.replace("．", ".").replace("，", ","))

    sqa = sq(sa)
    out = []
    for L in ocr_lines:
        ot = sq(str(L.get("text", "")))
        if not ot:
            continue
        if sqa in ot or ot in sqa:
            out.append(L)
    return out


def _union_bboxes(bboxes: list[list[int]]) -> list[int]:
    """合并多个 bbox 为最小外接矩形（union）。"""
    return [
        min(b[0] for b in bboxes),
        min(b[1] for b in bboxes),
        max(b[2] for b in bboxes),
        max(b[3] for b in bboxes),
    ]


def _try_merge_split_matches(actual: str, matches: list[dict]) -> list[int] | None:
    """
    当多行 OCR 文本是同一单元格内容换行拆分时（如冲击吸收能量 "120/124/124"
    被拆成 "120/124" 和 "/124" 两行），将所有行的 bbox 合并为一个大框。
    判断条件：matches 按 y 坐标排序后文本拼合（去空格）== sq(actual)。
    满足则返回合并 bbox，否则返回 None。
    """
    def sq(s: str) -> str:
        return re.sub(r"\s+", "", s.replace("．", ".").replace("，", ","))

    sqa = sq(clean_cell_text(str(actual).strip()))
    if not sqa or len(matches) < 2:
        return None
    sorted_m = sorted(matches, key=lambda L: (L["bbox"][1], L["bbox"][0]))
    combined = sq("".join(str(L.get("text", "")) for L in sorted_m))
    if combined == sqa:
        return _union_bboxes([L["bbox"] for L in sorted_m])
    return None


def _pick_ocr_bbox_for_failure(
    actual: str,
    cell_bbox: list[int] | None,
    ocr_lines: list[dict] | None,
    img_w: int | None = None,
    img_h: int | None = None,
    expand_px: int = 36,
) -> list[int] | None:
    """
    用「不合格实测值 actual」在 OCR 行中选框，与识别底图同源。
    多匹配时：扩边粗框内相交面积优先，其次惩罚过高 bbox（避免跨多行大块），再选面积较小、纵横向更近。
    """
    if not ocr_lines or not actual:
        return None
    matches = _ocr_lines_for_actual(actual, ocr_lines)
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]["bbox"]

    # 多行拼合检测：OCR 按换行拆成多行但拼合文本 = actual → 合并 bbox（如冲击吸收能量）
    merged = _try_merge_split_matches(actual, matches)
    if merged is not None:
        return merged

    ex = None
    coarse_h = 0.0
    if cell_bbox is not None and img_w is not None and img_h is not None:
        ex = _expand_bbox(cell_bbox, expand_px, img_w, img_h)
        coarse_h = float(cell_bbox[3] - cell_bbox[1])
    ccy = ccx = 0.0
    if cell_bbox is not None:
        ccy = (cell_bbox[1] + cell_bbox[3]) / 2
        ccx = (cell_bbox[0] + cell_bbox[2]) / 2

    sa = clean_cell_text(str(actual).strip())
    sqa = re.sub(r"\s+", "", sa.replace("．", ".").replace("，", ","))
    digitish = bool(sqa) and re.match(r"^[-+]?\d", sqa)

    def sq_ocr(L: dict) -> str:
        return re.sub(
            r"\s+", "",
            str(L.get("text", "")).replace("．", ".").replace("，", ","),
        )

    def sort_key(L: dict) -> tuple:
        bb = L["bbox"]
        if ex is not None:
            ia = _bbox_intersection_area(bb, ex)
        elif cell_bbox is not None:
            ia = _bbox_intersection_area(bb, cell_bbox)
        else:
            ia = 0
        cy = (bb[1] + bb[3]) / 2
        cx = (bb[0] + bb[2]) / 2
        bh = float(bb[3] - bb[1])
        area = float((bb[2] - bb[0]) * (bb[3] - bb[1]))
        tall_pen = 0.0
        if coarse_h > 6 and bh > coarse_h * 2.15:
            tall_pen = 1e6 + bh
        len_diff = abs(len(sq_ocr(L)) - len(sqa)) if digitish else 0
        return (ia, -tall_pen, -len_diff, -area, -abs(cy - ccy), -abs(cx - ccx))

    return max(matches, key=sort_key)["bbox"]


def export_ocr_mapping_excel(
    page_no: int,
    tbl_no: int,
    ws,
    tbl_info: dict,
    failures: list[dict],
    out_path: str,
) -> None:
    """
    输出 OCR 坐标映射调试 Excel，帮助核查识别位置是否准确。

    Sheet1「OCR文本行」  — 表格区域内所有 OCR 识别行，按 y→x 排序：
        序号 / 识别文本 / x1 / y1 / x2 / y2 / 宽 / 高

    Sheet2「单元格坐标」 — 每个非空单元格与其在图片上的匹配坐标：
        Excel行 / Excel列 / 列字母 / 单元格文本 / OCR_x1 / OCR_y1 / OCR_x2 / OCR_y2 / 宽 / 高

    Sheet3「不合格单元格」— 不合格项汇总：
        Excel行 / Excel列 / 实测值 / 判定原因 / OCR_x1 / OCR_y1 / OCR_x2 / OCR_y2
    """
    from openpyxl.styles import PatternFill as _PF

    wb_d = Workbook()
    BOLD = Font(bold=True)
    FILL_HDR = _PF(fill_type="solid", fgColor="D9E1F2")
    FILL_FAIL = _PF(fill_type="solid", fgColor="FFCCCC")

    def _set_header(sheet, headers: list[str]) -> None:
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = BOLD
            cell.fill = FILL_HDR

    def _auto_width(sheet) -> None:
        for col in sheet.columns:
            w = max((len(str(c.value or "")) for c in col), default=6)
            sheet.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

    # ── Sheet1：所有 OCR 文本行 ───────────────────────────────────────────────
    ws_ocr = wb_d.active
    ws_ocr.title = "OCR文本行"
    _set_header(ws_ocr, ["序号", "识别文本", "x1(左)", "y1(上)", "x2(右)", "y2(下)", "宽px", "高px"])
    ocr_lines = tbl_info.get("table_ocr_lines") or []
    for i, line in enumerate(sorted(ocr_lines, key=lambda L: (L["bbox"][1], L["bbox"][0])), 1):
        bb = line.get("bbox", [0, 0, 0, 0])
        x1, y1, x2, y2 = bb
        ws_ocr.append([i, line.get("text", ""), x1, y1, x2, y2, x2 - x1, y2 - y1])
    _auto_width(ws_ocr)

    # ── Sheet2：单元格坐标映射 ────────────────────────────────────────────────
    ws_map = wb_d.create_sheet("单元格坐标")
    _set_header(ws_map, ["Excel行", "Excel列", "列字母", "单元格文本",
                          "OCR_x1", "OCR_y1", "OCR_x2", "OCR_y2", "宽px", "高px"])
    rec_map = tbl_info.get("cell_recognition_boxes") or {}
    visited: set[tuple[int, int]] = set()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(row=r, column=c).value
            if raw is None or (r, c) in visited:
                continue
            visited.add((r, c))
            text = str(raw).strip()
            bb = rec_map.get((r, c))
            if bb:
                x1, y1, x2, y2 = bb
                ws_map.append([r, c, get_column_letter(c), text, x1, y1, x2, y2, x2 - x1, y2 - y1])
            else:
                ws_map.append([r, c, get_column_letter(c), text, "", "", "", "", "", ""])
    _auto_width(ws_map)

    # ── Sheet3：不合格单元格 ──────────────────────────────────────────────────
    ws_fail = wb_d.create_sheet("不合格单元格")
    _set_header(ws_fail, ["Excel行", "Excel列", "列字母", "实测值", "判定原因",
                           "OCR_x1", "OCR_y1", "OCR_x2", "OCR_y2"])
    for f in failures:
        r, c = f["excel_row"], f["excel_col"]
        bb = rec_map.get((r, c))
        row_data = [r, c, get_column_letter(c), f.get("actual", ""), f.get("reason", "")]
        if bb:
            row_data += list(bb)
        else:
            row_data += ["", "", "", ""]
        ws_fail.append(row_data)
        for cell in ws_fail[ws_fail.max_row]:
            cell.fill = FILL_FAIL
    _auto_width(ws_fail)

    wb_d.save(out_path)
    print(f"  [OCR映射] 已保存：{Path(out_path).name}"
          f"（OCR行 {len(ocr_lines)} 条，单元格 {len(visited)} 个）")


def mark_failures_on_image(
    img_path: str,
    failures_with_meta: list,
    out_path: str,
    td_coord_maps: dict | None = None,
    ocr_lines: list[dict] | None = None,
    tbl_ws_map: dict | None = None,
) -> None:
    """
    在 PDF 渲染图片上以红色矩形框标注所有不合格单元格，并在框旁标注原因。

    failures_with_meta: [(tbl_info_dict, excel_row, excel_col, actual, reason), ...]
    td_coord_maps: {tbl_info_id: td_coord_map}
    ocr_lines: 本页全文 OCR
    tbl_ws_map: {tbl_info_id: worksheet}，用于合并单元格感知网格粗定位（与 Excel 一致）
    策略：优先用「实测值 actual」在表内/全文 OCR 上匹配识别框 + 合并感知粗框消歧；
          再兜底 cell_recognition_boxes、最后 _get_cell_img_bbox。
    out_path: 保存标注后图片的路径（jpg）
    """
    img   = Image.open(img_path).convert("RGB")
    draw  = ImageDraw.Draw(img)
    drawn = 0
    W, H  = img.size

    for tbl_info, excel_row, excel_col, actual, reason in failures_with_meta:
        tbl_id       = id(tbl_info)
        td_coord_map = (td_coord_maps or {}).get(tbl_id, {})
        ws           = (tbl_ws_map or {}).get(tbl_id)
        if ws is not None:
            coarse = _coarse_bbox_from_ws_cell(ws, tbl_info, excel_row, excel_col)
        else:
            coarse = _grid_cell_bbox_only(tbl_info, excel_row, excel_col)

        # ── 混合定位策略：OCR x 轴（列精度高）+ coarse y 轴（行准确，无 OCR 偏移）──
        # OCR overall_ocr_res 存在系统性 y 轴偏低，但 x 轴列位置准确；
        # 均匀网格 coarse 行高准确但列宽估算偏左；两者互补得到精确单元格位置。
        scoped       = tbl_info.get("table_ocr_lines") or []
        prefer_lines = scoped if scoped else (ocr_lines or [])

        ocr_bbox  = _pick_ocr_bbox_for_failure(actual, coarse, prefer_lines, W, H)
        if ocr_bbox is None:
            ocr_bbox = _pick_ocr_bbox_for_failure(actual, coarse, ocr_lines, W, H)

        if ocr_bbox is not None and coarse is not None:
            # OCR x 中心（列定位）+ coarse 宽度（单元格宽）+ coarse y（行定位）
            # 以 OCR 文字中心为水平中心，避免数字偏边导致框偏移
            ocr_cx  = (ocr_bbox[0] + ocr_bbox[2]) / 2.0
            cell_w  = coarse[2] - coarse[0]
            x1 = max(0, int(round(ocr_cx - cell_w / 2)))
            x2 = min(W, int(round(ocr_cx + cell_w / 2)))
            draw_bbox = [x1, coarse[1], x2, coarse[3]]
        elif ocr_bbox is not None:
            draw_bbox = ocr_bbox
        else:
            # OCR 匹配失败：退回精确 cell_bboxes 或均匀网格
            draw_bbox = _get_cell_img_bbox(tbl_info, td_coord_map, excel_row, excel_col)
            if draw_bbox is None:
                draw_bbox = coarse

        if draw_bbox is None:
            print(f"  [跳过] 行{excel_row} 列{excel_col} 无法获取图片坐标，跳过标注。")
            continue

        ix1, iy1, ix2, iy2 = draw_bbox
        draw.rectangle([ix1, iy1, ix2, iy2], outline=(220, 50, 50), width=3)

        label  = reason or actual
        text_x = min(ix2 + 4, img.width - 1)
        text_y = max(iy1, 0)
        draw.text((text_x, text_y), label, fill=(220, 50, 50))
        drawn += 1

    if drawn > 0:
        img.save(out_path, quality=95)
        print(f"  标注图片已保存：{Path(out_path).name}（共标注 {drawn} 处）")
    else:
        print(f"  [跳过] 未画出任何标注框，图片未保存。")


# ─── 主流程 ───────────────────────────────────────────────────────────────────

def ocr_pdf_ppstruct(
    pdf_path: str,
    output_dir: str,
    report_type: str = "力学检测",
    remove_stamp: bool = True,
):
    """
    report_type:   报告类型，用于输出文件命名前缀（"力学检测"/"化学检测" 等）
    remove_stamp:  True 时自动去除图片/PDF 中的红色公章，减少 OCR 干扰
    输出文件命名规则：
      识别底图_{页}_{时间戳}.png                   — 与引擎共用底图（已去章，坐标同源）
      错误标注_{编号}_{时间戳}.jpg                 — 不合格标注图
      表格_{report_type}_{时间戳}.xlsx             — 识别结果表格
    """
    pdf_name    = Path(pdf_path).stem
    total_start = time.perf_counter()
    timestamp   = datetime.now().strftime("%Y%m%d%H%M%S")

    # 在指定输出目录下，以"文件名_识别结果_时间戳"新建子文件夹
    out_dir = Path(output_dir) / f"{pdf_name}_识别结果_{timestamp}"
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'═'*55}")
    print(f"  [PP-Structure V3]  {Path(pdf_path).name}")
    print(f"  报告类型：{report_type}")
    print(f"  输出目录：{out_dir}")
    print(f"  时间戳：{timestamp}")
    print(f"{'═'*55}")

    # ── 1. 准备图片（PDF 渲染 或 直接使用图片）────────────────────────────────
    # 临时目录：存放渲染中间图片，识别完成后自动删除
    tmp_dir = out_dir / "_tmp"
    suffix  = Path(pdf_path).suffix.lower()
    if suffix == ".pdf":
        print(f"\n[1/4] 渲染 PDF 页面{'（自动去章）' if remove_stamp else ''}…")
        pages = pdf_to_images(pdf_path, tmp_dir=tmp_dir, remove_stamp=remove_stamp)
    elif suffix in (".jpg", ".jpeg", ".png"):
        print(f"\n[1/4] 读取图片文件{'（自动去章）' if remove_stamp else ''}…")
        pages = image_to_pages(pdf_path, tmp_dir=tmp_dir, remove_stamp=remove_stamp)
    else:
        print(f"  [错误] 不支持的文件格式：{suffix}，仅支持 PDF / JPG / JPEG / PNG")
        return

    # 将临时渲染图复制到输出目录：后续 predict 与错误标注共用此路径，坐标与像素严格一致
    archived_pages: list[tuple[int, str]] = []
    for page_no, pth in pages:
        dest = out_dir / f"识别底图_{page_no}_{timestamp}.png"
        shutil.copy2(pth, dest)
        archived_pages.append((page_no, str(dest)))
    pages = archived_pages
    print(f"  已保存识别底图至输出目录（共 {len(pages)} 张，与 OCR 坐标系一致）")

    page_sizes: dict[int, tuple[int, int]] = {}
    for page_no, pth in pages:
        with Image.open(pth) as im:
            page_sizes[page_no] = im.size

    # ── 2. PPStructureV3 识别 ─────────────────────────────────────────────────
    print("\n[2/4] 启动 PPStructureV3…")
    engine = PPStructureV3(lang="ch", device="cpu")

    wb  = Workbook()
    # all_tables: [(页码, 表格序号, tbl_info_dict)]
    all_tables: list[tuple[int, int, dict]] = []
    # 每页 OCR 文本行（与 predict 同源），用于错误标注时贴合原图识别位置
    page_ocr_lines: dict[int, list[dict]] = {}

    for page_no, img_path in pages:
        t0 = time.perf_counter()
        print(f"\n  第 {page_no} 页  识别中…")
        tables, ocr_lines = extract_tables_from_page(engine, img_path)
        page_ocr_lines[page_no] = ocr_lines
        elapsed = time.perf_counter() - t0
        print(f"  第 {page_no} 页  检测到 {len(tables)} 个表格  耗时 {elapsed:.1f}s")

        for j, tbl_info in enumerate(tables, 1):
            all_tables.append((page_no, j, tbl_info))
            # 保存原始 HTML 供调试
            html_path = out_dir / f"{pdf_name}_ppstruct_page_{page_no}_table_{j}.html"
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(f"<html><meta charset='utf-8'><body>{tbl_info['html']}</body></html>")
            print(f"    → HTML 已保存：{html_path.name}")

    # ── 3. 写入 Excel ─────────────────────────────────────────────────────────
    print(f"\n[3/4] 写入 Excel…  共 {len(all_tables)} 个表格")

    if not all_tables:
        print("  未检测到任何表格，退出。")
        return

    # page_failures_map: page_no → [(tbl_info, excel_row, excel_col, actual, reason)]
    page_failures_map: dict[int, list] = defaultdict(list)
    # td_coord_maps: tbl_info id → td_coord_map（用于精确图片标注）
    td_coord_maps: dict[int, dict] = {}
    # 错误标注时按合并区域算网格，需 worksheet 引用
    tbl_ws_map: dict[int, object] = {}
    xlsx_path = out_dir / f"表格_{report_type}_{timestamp}.xlsx"
    db_rows: list[dict] = []

    ensure_nonconformity_table()

    for idx, (page_no, tbl_no, tbl_info) in enumerate(all_tables):
        sheet_name = f"P{page_no}_T{tbl_no}" if len(all_tables) > 1 else "Sheet1"
        ws = wb.active if idx == 0 else wb.create_sheet(sheet_name)
        if idx == 0:
            ws.title = sheet_name
        last_row, td_coord_map = html_table_to_sheet(ws, tbl_info["html"], row_offset=1)
        tid = id(tbl_info)
        td_coord_maps[tid] = td_coord_map
        tbl_ws_map[tid] = ws
        build_cell_recognition_boxes(
            ws, tbl_info, td_coord_map, page_sizes[page_no]
        )
        max_col  = ws.max_column
        apply_header_style(ws, last_row, max_col)
        failures = apply_value_judgment(ws)
        auto_column_width(ws)
        print(f"  Sheet [{sheet_name}]  {last_row-1} 行 × {max_col} 列")

        # 导出 OCR 坐标映射调试 Excel
        ocr_map_path = out_dir / f"OCR坐标映射_P{page_no}_T{tbl_no}_{timestamp}.xlsx"
        export_ocr_mapping_excel(page_no, tbl_no, ws, tbl_info, failures, str(ocr_map_path))

        for f in failures:
            page_failures_map[page_no].append(
                (tbl_info, f["excel_row"], f["excel_col"], f["actual"], f["reason"])
            )
            db_rows.append(
                build_record(
                    pdf_path=pdf_path,
                    report_type=report_type,
                    xlsx_path=str(xlsx_path),
                    batch_id=timestamp,
                    sheet_name=sheet_name,
                    page_no=page_no,
                    table_index=tbl_no,
                    excel_row=f["excel_row"],
                    excel_col=f["excel_col"],
                    sample_no=f["sample_no"],
                    test_item=f["test_item"],
                    standard_text=f["standard_text"],
                    rule_type=f["rule_type"],
                    actual_value=f["actual"],
                    fail_reason=f["reason"],
                    col_letter=get_column_letter(f["excel_col"]),
                )
            )

    wb.save(str(xlsx_path))
    insert_nonconformity_records(db_rows)

    # ── 4. 生成不合格标注图片 ─────────────────────────────────────────────────
    print(f"\n[4/4] 生成不合格标注图片…")
    marked_count = 0
    for page_no, img_path in pages:
        failures_this_page = page_failures_map.get(page_no, [])
        if not failures_this_page:
            print(f"  第 {page_no} 页  无不合格项，跳过图片标注。")
            continue
        marked_path = out_dir / f"错误标注_{page_no}_{timestamp}.jpg"
        mark_failures_on_image(
            img_path,
            failures_this_page,
            str(marked_path),
            td_coord_maps,
            ocr_lines=page_ocr_lines.get(page_no),
            tbl_ws_map=tbl_ws_map,
        )
        marked_count += 1

    # 清理临时图片目录（识别底图已复制到 out_dir）
    if tmp_dir.exists():
        shutil.rmtree(tmp_dir)
        print("  [清理] 临时图片已删除。")

    total_elapsed = time.perf_counter() - total_start
    print(f"\n{'═'*55}")
    print(f"✓ 完成！总耗时：{total_elapsed:.1f}s")
    print(f"  {xlsx_path.name}  — 识别结果表格（含合规标注）")
    if marked_count:
        print(f"  错误标注_N_{timestamp}.jpg  — 不合格标注图片（共 {marked_count} 页）")
    print(f"{'═'*55}")


# ─── stdout 重定向 ────────────────────────────────────────────────────────────

import queue
import subprocess
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk


class _QueueWriter:
    """将 stdout/stderr 写入线程安全队列，供 GUI 主线程消费。"""

    def __init__(self, q: queue.Queue):
        self._q = q

    def write(self, text: str):
        if text:
            self._q.put(text)

    def flush(self):
        pass


# ─── GUI 主界面 ───────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF 表格识别工具（PP-Structure V3）")
        self.minsize(720, 560)
        self.resizable(True, True)

        self._output_dir    = tk.StringVar(value="/Users/project/QMS/PaddleOCR/QMS/output2")
        self._remove_stamp  = tk.BooleanVar(value=True)
        self._pdf_files: list[str] = [
            "/Users/project/QMS/PaddleOCR/QMS/test_files/机械性能复检报告.pdf"
        ]
        self._log_queue: queue.Queue = queue.Queue()
        self._running = False

        self._build_ui()
        self._poll_log()
        self._refresh_files_text()

    # ─── UI 构建 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        # ── 输入文件区 ──
        frm_in = ttk.LabelFrame(self, text="输入 PDF 文件", padding=8)
        frm_in.pack(fill="x", **pad)

        self._files_text = tk.Text(
            frm_in, height=4, state="disabled",
            wrap="none", relief="sunken", bd=1,
            font=("Menlo", 12),
        )
        self._files_text.pack(side="left", fill="both", expand=True, padx=(0, 8))

        frm_in_btns = ttk.Frame(frm_in)
        frm_in_btns.pack(side="right", fill="y")
        ttk.Button(frm_in_btns, text="选择文件…", command=self._pick_files, width=10).pack(
            fill="x", pady=(0, 4))
        ttk.Button(frm_in_btns, text="清  空", command=self._clear_files, width=10).pack(fill="x")

        # ── 输出目录区 ──
        frm_out = ttk.LabelFrame(self, text="输出目录", padding=8)
        frm_out.pack(fill="x", **pad)

        ttk.Entry(frm_out, textvariable=self._output_dir, font=("Menlo", 12)).pack(
            side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(frm_out, text="选择目录…", command=self._pick_output).pack(side="left", padx=(0, 4))
        ttk.Button(frm_out, text="打开目录", command=self._open_output).pack(side="left")

        # ── 操作按钮与状态 ──
        frm_btn = ttk.Frame(self)
        frm_btn.pack(fill="x", padx=12, pady=(8, 2))

        self._btn_start = ttk.Button(frm_btn, text="▶  开始识别", command=self._start, width=14)
        self._btn_start.pack(side="left", padx=(0, 8))

        self._btn_stop = ttk.Button(frm_btn, text="■  停止", command=self._stop,
                                    state="disabled", width=10)
        self._btn_stop.pack(side="left", padx=(0, 12))

        ttk.Checkbutton(
            frm_btn, text="自动去除公章", variable=self._remove_stamp,
        ).pack(side="left", padx=(0, 12))

        self._status_var = tk.StringVar(value="就绪")
        ttk.Label(frm_btn, textvariable=self._status_var, foreground="gray").pack(side="left")

        # ── 进度条 ──
        self._progress = ttk.Progressbar(self, mode="indeterminate")
        self._progress.pack(fill="x", padx=12, pady=(2, 4))

        # ── 日志区 ──
        frm_log = ttk.LabelFrame(self, text="运行日志", padding=8)
        frm_log.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self._log = scrolledtext.ScrolledText(
            frm_log, state="disabled",
            font=("Menlo", 11), wrap="none",
            relief="flat", bd=0,
        )
        self._log.pack(fill="both", expand=True)

    # ─── 文件 / 目录选择 ──────────────────────────────────────────────────────

    def _pick_files(self):
        files = filedialog.askopenfilenames(
            title="选择 PDF 或图片文件",
            filetypes=[
                ("支持的格式", "*.pdf *.jpg *.jpeg *.png"),
                ("PDF 文件", "*.pdf"),
                ("图片文件", "*.jpg *.jpeg *.png"),
                ("所有文件", "*.*"),
            ],
        )
        if files:
            self._pdf_files = list(files)
            self._refresh_files_text()

    def _clear_files(self):
        self._pdf_files = []
        self._refresh_files_text()

    def _refresh_files_text(self):
        self._files_text.config(state="normal")
        self._files_text.delete("1.0", "end")
        for f in self._pdf_files:
            self._files_text.insert("end", f + "\n")
        self._files_text.config(state="disabled")

    def _pick_output(self):
        d = filedialog.askdirectory(title="选择输出目录")
        if d:
            self._output_dir.set(d)

    def _open_output(self):
        raw  = self._output_dir.get().strip()
        path = Path(raw) if Path(raw).is_absolute() else Path(__file__).parent / raw
        path.mkdir(parents=True, exist_ok=True)
        subprocess.Popen(["open", str(path)])

    # ─── 识别控制 ─────────────────────────────────────────────────────────────

    def _start(self):
        if not self._pdf_files:
            messagebox.showwarning("提示", "请先选择至少一个 PDF 文件。")
            return
        if not self._output_dir.get().strip():
            messagebox.showwarning("提示", "请选择输出目录。")
            return

        self._running = True
        self._btn_start.config(state="disabled")
        self._btn_stop.config(state="normal")
        self._progress.start(12)
        self._status_var.set("识别中…")
        self._log_clear()

        threading.Thread(target=self._run_ocr, daemon=True).start()

    def _stop(self):
        self._running = False
        self._btn_stop.config(state="disabled")
        self._status_var.set("正在等待当前文件完成后停止…")

    def _run_ocr(self):
        old_stdout, old_stderr = sys.stdout, sys.stderr
        writer = _QueueWriter(self._log_queue)
        sys.stdout = writer
        sys.stderr = writer

        try:
            total     = len(self._pdf_files)
            completed = 0
            for idx, pdf_path in enumerate(self._pdf_files, 1):
                if not self._running:
                    print("\n[停止] 已中止识别。")
                    break
                print(f"\n{'═'*50}")
                print(f"[{idx}/{total}]  {Path(pdf_path).name}")
                print(f"{'═'*50}")
                try:
                    ocr_pdf_ppstruct(
                        pdf_path,
                        self._output_dir.get().strip(),
                        remove_stamp=self._remove_stamp.get(),
                    )
                    completed += 1
                except Exception as exc:
                    print(f"\n[错误] 处理失败：{exc}")

            if self._running:
                print(f"\n{'═'*50}")
                print(f"✓ 全部完成：{completed}/{total} 个文件处理成功。")
                print(f"{'═'*50}")
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr
            self.after(0, self._on_done)

    def _on_done(self):
        self._running = False
        self._progress.stop()
        self._btn_start.config(state="normal")
        self._btn_stop.config(state="disabled")
        self._status_var.set("完成")

    # ─── 日志轮询 ─────────────────────────────────────────────────────────────

    def _log_clear(self):
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")

    def _poll_log(self):
        """每 100 ms 从队列取出日志文本，写入控件（主线程安全）。"""
        try:
            while True:
                text = self._log_queue.get_nowait()
                self._log.config(state="normal")
                self._log.insert("end", text)
                self._log.see("end")
                self._log.config(state="disabled")
        except queue.Empty:
            pass
        self.after(100, self._poll_log)


# ─── 入口 ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if "-i" in sys.argv or "--input" in sys.argv:
        args = parse_args()
        ocr_pdf_ppstruct(args.input, args.output)
    else:
        App().mainloop()
