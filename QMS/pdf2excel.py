"""
PDF 表格识别 → 三文件输出工具

使用 PP-OCRv5 识别 PDF 中的印刷体文字，每个 PDF 产生三类输出：

  1. <pdf名>_table.xlsx          — 还原格式可编辑表格（行列聚类 + 单元格合并）
  2. <pdf名>_data.xlsx           — 数据存储表格（含行列号、低置信度标红）
  3. <pdf名>_page_N_res.json     — 增强 JSON（含 row_nums / col_nums）

每页额外输出两张标注图：
  · _page_N_low_conf.jpg         — 原有：低置信度文本框（红色线框）
  · _page_N_conf_heat.jpg        — 新增：置信度热力图（仅低置信度区域，
                                    差距越大红色越深，半透明填充保留原文）

用法：
    conda activate paddleocr
    python pdf2excel.py -i <PDF路径> [-o <输出目录>] [--threshold 0.991] [--row-tol 15]

示例：
    python pdf2excel.py -i "/Users/project/QMS/PaddleOCR/文件/机械性能复检报告.pdf"
    python pdf2excel.py -i "/Users/project/QMS/PaddleOCR/文件/机械性能复检报告.pdf" -o /tmp/out --threshold 0.95

默认参数从同级 config.toml 读取；命令行参数优先级高于配置文件。
运行 `python config_ui.py` 可打开可视化配置页面。
"""

import argparse
import json
import os
import string
import time
from pathlib import Path

os.environ["PADDLE_PDX_MODEL_SOURCE"] = "modelscope"
os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

import numpy as np
import pypdfium2 as pdfium
from PIL import Image, ImageDraw
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from paddleocr import PaddleOCR

from ocr_config import load_config

try:
    import cv2 as _cv2
    import numpy as _np
    _CV2_AVAILABLE = True
except ImportError:
    _CV2_AVAILABLE = False
    _np = None

# ─── 常量 ─────────────────────────────────────────────────────────────────────

MAX_IMAGE_WIDTH    = 1600   # 提高分辨率可改善 ☑/□ 等小符号的识别率
LOW_CONF_BOX_COLOR = "red"
LOW_CONF_BOX_WIDTH = 3
LOW_CONF_FILL_COLOR = "FFCCCC"
COL_GAP          = 10   # 列边界聚类间距（px）：x_left 差值 ≤ COL_GAP 归为同一列
COL_MERGE_ADVANCE = 25  # 跨列合并最低入侵阈值（px）：x_right 必须超过下一列左边界此距离才合并


# ─── 命令行参数 ───────────────────────────────────────────────────────────────

def parse_args():
    cfg = load_config()
    parser = argparse.ArgumentParser(
        description="PDF 表格识别并导出三文件",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("-i", "--input", required=True, help="PDF 文件路径")
    parser.add_argument(
        "-o", "--output",
        default=cfg["output"]["output_dir"],
        help=f"输出目录（config.toml 默认：{cfg['output']['output_dir']}）",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=cfg["ocr"]["threshold"],
        help="置信度阈值，低于此值将标红",
    )
    parser.add_argument(
        "--row-tol",
        type=int,
        default=cfg["ocr"]["row_tol"],
        help="行分组 Y 差值容忍度（px）",
    )
    return parser.parse_args()


# ─── PDF 渲染 ─────────────────────────────────────────────────────────────────

def pdf_to_images(pdf_path: str, max_width: int = MAX_IMAGE_WIDTH) -> list:
    """将 PDF 每页渲染为 PIL Image，限制最大宽度避免内存溢出。"""
    pdf = pdfium.PdfDocument(pdf_path)
    images = []
    for i in range(len(pdf)):
        page = pdf[i]
        width, height = page.get_size()
        scale = min(max_width / width, 2.0)
        bitmap = page.render(scale=scale)
        img = bitmap.to_pil()
        images.append(img)
        print(f"  第 {i+1} 页 → {img.size[0]}×{img.size[1]}")
    return images


# ─── 低置信度标注图 ───────────────────────────────────────────────────────────

def draw_low_conf_boxes(img_path: str, low_conf_items: list, out_path: str):
    """在渲染图上用红色线框标注低置信度文本块（原有逻辑保留）。"""
    if not low_conf_items:
        return
    img = Image.open(img_path).convert("RGB")
    draw = ImageDraw.Draw(img)
    for _text, _score, poly in low_conf_items:
        pts = [(int(p[0]), int(p[1])) for p in poly]
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        x0, y0, x1, y1 = min(xs), min(ys), max(xs), max(ys)
        for w in range(LOW_CONF_BOX_WIDTH):
            draw.rectangle([x0 - w, y0 - w, x1 + w, y1 + w], outline=LOW_CONF_BOX_COLOR)
    img.save(out_path, quality=95)
    print(f"  低置信度标注图 → {out_path}")


# ─── 打勾检测（opencv-python 可用时启用）────────────────────────────────────────

# □ 被识别为空方框时，用此函数重新判断是否已打勾
_CHECKBOX_CHARS = {"□", "■"}
_CHECKED_CHAR   = "☑"
_UNCHECKED_CHAR = "□"

def _is_checked_by_pixels(img_path: str, poly: list, fill_ratio_threshold: float = 0.20) -> bool:
    """
    只裁剪 poly 左端「checkbox 字符」所在的正方形小区域（高度 × 高度），
    统计深色像素占比来判断是否已打勾。

    之所以只取左端正方形：□/☑ 字符宽约等于行高，后续文字不应纳入判断。
    fill_ratio_threshold: 深色像素（灰度 < 100）占比阈值，默认 20%。
    """
    if not _CV2_AVAILABLE or len(poly) == 0:
        return False
    try:
        img = _cv2.imread(img_path, _cv2.IMREAD_GRAYSCALE)
        if img is None:
            return False
        xs = [int(p[0]) for p in poly]
        ys = [int(p[1]) for p in poly]
        x0, y0 = max(0, min(xs)), max(0, min(ys))
        x1, y1 = min(img.shape[1], max(xs)), min(img.shape[0], max(ys))
        h = y1 - y0
        if h <= 2 or x1 <= x0:
            return False

        # 只取左端一个字符宽度（约等于行高），并向内收缩 15% 边距排除边框像素
        char_w  = min(h, x1 - x0)
        pad     = max(1, int(char_w * 0.15))
        cx0     = x0 + pad
        cx1     = x0 + char_w - pad
        cy0     = y0 + pad
        cy1     = y1 - pad
        if cx1 <= cx0 or cy1 <= cy0:
            return False

        crop        = img[cy0:cy1, cx0:cx1]
        dark_pixels = int((crop < 100).sum())   # 深色像素（更严格）
        total       = crop.size
        return (dark_pixels / total) > fill_ratio_threshold
    except Exception:
        return False


def _has_checkbox_at_left(img_path: str, poly: list,
                           char_h_ratio: float = 1.1) -> tuple[bool, bool]:
    """
    在 poly 左端检测是否存在 checkbox 方框，以及是否已打勾。

    返回 (has_checkbox: bool, is_checked: bool)

    算法：
    1. 取文本框高度 h，在左端截取 h×h 的正方形区域
    2. 检测方框边缘（中空矩形）→ 确认有 checkbox
    3. 在方框内部统计深色像素 → 判断是否已打勾
    """
    if not _CV2_AVAILABLE or len(poly) == 0:
        return False, False
    try:
        gray = _cv2.imread(img_path, _cv2.IMREAD_GRAYSCALE)
        if gray is None:
            return False, False

        xs = [int(p[0]) for p in poly]
        ys = [int(p[1]) for p in poly]
        x0, y0 = max(0, min(xs)), max(0, min(ys))
        x1, y1 = min(gray.shape[1], max(xs)), min(gray.shape[0], max(ys))
        h = y1 - y0
        if h < 6:
            return False, False

        # 截取左端正方形
        sq_w  = min(int(h * char_h_ratio), x1 - x0)
        sq    = gray[y0: y1, x0: x0 + sq_w]
        _, bw = _cv2.threshold(sq, 0, 255, _cv2.THRESH_BINARY_INV + _cv2.THRESH_OTSU)

        # 检测是否有方框：边缘区域（外 20%）像素密度高
        outer_mask = _np.zeros_like(bw)
        pad = max(1, int(min(sq.shape) * 0.15))
        outer_mask[:pad, :]  = 255
        outer_mask[-pad:, :] = 255
        outer_mask[:, :pad]  = 255
        outer_mask[:, -pad:] = 255
        border_ratio = float((bw & outer_mask).sum()) / (255 * outer_mask.sum() / 255 + 1e-6)
        has_checkbox = border_ratio > 0.25   # 边框像素占比 > 25%

        if not has_checkbox:
            return False, False

        # 检测内部是否有勾：内部区域深色像素密度
        inner = bw[pad:-pad, pad:-pad]
        if inner.size == 0:
            return True, False
        fill_ratio = float(inner.sum()) / (255 * inner.size)
        is_checked = fill_ratio > 0.12   # 内部深色像素 > 12%

        return True, is_checked
    except Exception:
        return False, False


def detect_checkboxes(items: list, img_path: str) -> list:
    """
    扫描每条 item 的文本框左侧，主动检测 checkbox 是否存在及是否已打勾：

    · 如果 OCR 已输出 □/■ 前缀  → 用像素结果修正（□→☑ 或保持 □）
    · 如果 OCR 没有输出前缀但检测到 checkbox → 补加 ☑ 或 □ 前缀
    · 没有 checkbox 区域 → text 不变

    items    : list of dict，含 text / poly 字段（in-place 修改）
    img_path : 该页渲染图路径
    """
    if not _CV2_AVAILABLE:
        return items

    for item in items:
        text = item.get("text", "")
        poly = item.get("poly", [])
        if not text or len(poly) == 0:
            continue

        has_cb, is_checked = _has_checkbox_at_left(img_path, poly)

        if text[0] in _CHECKBOX_CHARS:
            # OCR 已识别出 □，用像素结果校正
            if has_cb:
                item["text"] = (_CHECKED_CHAR if is_checked else _UNCHECKED_CHAR) + text[1:]
                item["checkbox_checked"] = is_checked
        elif has_cb:
            # OCR 漏了 checkbox 前缀，补加
            item["text"] = (_CHECKED_CHAR if is_checked else _UNCHECKED_CHAR) + text
            item["checkbox_checked"] = is_checked

    return items


def _heat_color(score: float, threshold: float) -> tuple:
    """
    根据置信度与阈值的差距，返回 RGBA 热力颜色。

    差距为 0（刚好在阈值边缘）→ 浅粉色半透明
    差距最大（score 趋近 0）   → 深红色不透明

    颜色范围（t = 0~1）：
      R : 255  → 200   （固定高红通道）
      G : 230  →   0   （绿通道随差距归零）
      B : 230  →   0   （蓝通道随差距归零）
      A : 110  → 230   （越差越不透明）
    """
    t = min(1.0, (threshold - score) / max(threshold, 1e-9))
    r = int(255 - 55  * t)
    g = int(230 * (1.0 - t))
    b = int(230 * (1.0 - t))
    a = int(110 + 120 * t)
    return (r, g, b, a)


def draw_conf_heatmap(
    img_path: str,
    low_conf_items: list,
    threshold: float,
    out_path: str,
):
    """
    生成置信度热力图：仅标注低置信度区域，差距越大红色越深。

    · 半透明填充覆盖原图，仍能透过看到原始文字
    · 边框比填充颜色更深，突出边界
    · 图片右上角附带图例（浅→深 色阶条 + 文字说明）

    low_conf_items: list of (text, score, poly)
    """
    if not low_conf_items:
        return

    base    = Image.open(img_path).convert("RGBA")
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw    = ImageDraw.Draw(overlay)

    for _text, score, poly in low_conf_items:
        pts = [(int(p[0]), int(p[1])) for p in poly]
        xs  = [p[0] for p in pts]
        ys  = [p[1] for p in pts]
        x0, y0, x1, y1 = min(xs), min(ys), max(xs), max(ys)

        fill    = _heat_color(score, threshold)
        # 边框：R/G/B 各暗 40，alpha 全不透明
        outline = (max(0, fill[0] - 40), max(0, fill[1] - 40), max(0, fill[2] - 40), 255)

        draw.rectangle([x0, y0, x1, y1], fill=fill, outline=outline, width=2)

    # ── 图例：右上角绘制渐变色阶条
    _draw_legend(draw, base.size, threshold)

    composite = Image.alpha_composite(base, overlay).convert("RGB")
    composite.save(out_path, quality=95)
    print(f"  置信度热力图   → {out_path}")


def _draw_legend(draw: ImageDraw.ImageDraw, img_size: tuple, threshold: float):
    """在 overlay 上绘制右上角图例（渐变色阶条 + 文字）。"""
    W, _H   = img_size
    bar_w   = 16          # 色阶条宽度（px）
    bar_h   = 80          # 色阶条高度（px）
    margin  = 10
    x_right = W - margin
    y_top   = margin

    # 渐变色阶条（从浅到深，共 bar_h 步）
    for step in range(bar_h):
        t    = step / max(bar_h - 1, 1)
        r    = int(255 - 55  * t)
        g    = int(230 * (1.0 - t))
        b    = int(230 * (1.0 - t))
        a    = int(110 + 120 * t)
        y_px = y_top + step
        draw.rectangle(
            [x_right - bar_w, y_px, x_right, y_px + 1],
            fill=(r, g, b, a),
        )

    # 色阶条边框
    draw.rectangle(
        [x_right - bar_w, y_top, x_right, y_top + bar_h],
        outline=(100, 0, 0, 255),
        width=1,
    )

    # 标注文字（上：阈值，下：0）
    label_top    = f"≈{threshold:.3f}"
    label_bottom = "0.000"
    draw.text((x_right - bar_w - 2, y_top),           label_top,    fill=(80, 0, 0, 220), anchor="rm")
    draw.text((x_right - bar_w - 2, y_top + bar_h),   label_bottom, fill=(80, 0, 0, 220), anchor="rm")


# ─── 行列聚类 ─────────────────────────────────────────────────────────────────

def col_num_to_letter(n: int) -> str:
    """1-based 列号转字母：1→A, 26→Z, 27→AA …"""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = string.ascii_uppercase[rem] + result
    return result


def letter_to_col_idx(letters: str) -> int:
    """字母列名转 1-based 整数：A→1, Z→26, AA→27 …"""
    result = 0
    for ch in letters.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def assign_row_col(items: list, row_tol: int = 15) -> tuple[list, list]:
    """
    对一页的 OCR 识别结果进行行列聚类，并检测单元格合并范围。

    每条 item（dict）新增字段：
      row_num     : int   — 起始行号（1-based）
      row_num_end : int   — 终止行号（跨行时 > row_num，否则 == row_num）
      col_num     : str   — 起始列字母（A, B, C …）
      col_num_end : str   — 终止列字母（跨列时 > col_num，否则 == col_num）

    返回：(enriched_items, col_boundaries)
      col_boundaries: list of (x_start, x_end) — 每列 x 范围，用于设置列宽
    """
    if not items:
        return items, []

    # ── Step 1：提取各顶点坐标
    for item in items:
        poly = item.get("poly", [])
        if len(poly) >= 3:
            xs = [p[0] for p in poly]
            ys = [p[1] for p in poly]
            item["_x_left"]   = float(min(xs))
            item["_x_right"]  = float(max(xs))
            item["_y_top"]    = float(min(ys))
            item["_y_bottom"] = float(max(ys))
            item["_y_center"] = (item["_y_top"] + item["_y_bottom"]) / 2.0
        else:
            item["_x_left"] = item["_x_right"] = 0.0
            item["_y_top"]  = item["_y_bottom"] = item["_y_center"] = 0.0

    # ── Step 2：Y 轴聚类 → 行分组
    sorted_items = sorted(items, key=lambda x: x["_y_center"])
    rows: list[list] = []
    current_row = [sorted_items[0]]
    current_y   = sorted_items[0]["_y_center"]

    for item in sorted_items[1:]:
        if abs(item["_y_center"] - current_y) <= row_tol:
            current_row.append(item)
            current_y = sum(i["_y_center"] for i in current_row) / len(current_row)
        else:
            rows.append(current_row)
            current_row = [item]
            current_y   = item["_y_center"]
    rows.append(current_row)

    # 赋值行号，记录每行 y 中心
    row_y_centers: list[float] = []
    for row_num, row in enumerate(rows, 1):
        for item in row:
            item["row_num"] = row_num
        row_y_centers.append(sum(i["_y_center"] for i in row) / len(row))

    # ── Step 3：X 轴全局列对齐
    #
    # col_starts : 每列 x_left 聚类起点列表（仅用于列归属判断，不混入 x_right）
    # col_display_widths : [[x_left_min, x_right_max], ...] 仅用于 Excel 列宽估算
    #
    # 分开两个结构是关键：若把 x_right 写入同一个 boundary，
    # 宽标题行会把 col_boundaries[0][1] 扩到页面全宽，
    # 导致 find_col_start 把所有文本都归到 column A。
    x_lefts = sorted(item["_x_left"] for item in items)
    col_starts: list[float] = []     # 每列代表性 x_left（聚类后的区间中点）
    col_clusters: list[list[float]] = []  # 每列实际包含的所有 x_left 值

    if x_lefts:
        g_start = g_end = x_lefts[0]
        cluster: list[float] = [x_lefts[0]]
        for x in x_lefts[1:]:
            if x - g_end <= COL_GAP:
                g_end = x
                cluster.append(x)
            else:
                col_starts.append((g_start + g_end) / 2.0)
                col_clusters.append(cluster)
                g_start = g_end = x
                cluster = [x]
        col_starts.append((g_start + g_end) / 2.0)
        col_clusters.append(cluster)

    # 每列显示宽度：x_left_min ~ max(x_right) of all items in that column
    col_display_widths: list[list[float]] = [
        [min(cl), max(cl)] for cl in col_clusters
    ]
    for item in items:
        xl = item["_x_left"]
        # 找最近的列
        if col_starts:
            best = min(range(len(col_starts)), key=lambda i: abs(col_starts[i] - xl))
            col_display_widths[best][1] = max(col_display_widths[best][1], item["_x_right"])

    def find_col_start(x_left: float) -> int:
        """
        x_left 落入的列号（1-based）。

        策略：从右往左，找最后一个 col_start <= x_left + COL_GAP 的列。
        这样避免了宽项目扩展右边界后把所有内容归到第一列的问题。
        """
        if not col_starts:
            return 1
        for idx in range(len(col_starts) - 1, -1, -1):
            if col_starts[idx] <= x_left + COL_GAP:
                return idx + 1
        return 1

    def find_col_end(x_right: float, col_start_idx: int) -> int:
        """
        x_right 覆盖到的最后一列。

        只有 x_right 超过某列起点 COL_MERGE_ADVANCE px 以上才认为跨到该列。
        从 col_start_idx 的下一列开始向右检测，遇不满足立即停止。
        """
        result = col_start_idx
        for idx in range(col_start_idx, len(col_starts)):   # 0-based，对应列号 idx+1
            if col_starts[idx] + COL_MERGE_ADVANCE <= x_right:
                result = idx + 1
            else:
                break
        return result

    def find_row_end(y_bottom: float, row_start: int) -> int:
        """
        y_bottom 覆盖到的最后一行（行 y 中心 ≤ y_bottom + row_tol 的最大行号）。
        从 row_start 的下一行开始向下检测，遇到不满足条件立即停止。
        """
        result = row_start
        for r_idx in range(row_start, len(row_y_centers)):   # r_idx 是 0-based，对应行号 r_idx+1
            if row_y_centers[r_idx] <= y_bottom + row_tol:
                result = r_idx + 1
            else:
                break
        return result

    # ── Step 4：赋值列号及合并范围
    for item in items:
        c_start = find_col_start(item["_x_left"])
        c_end   = find_col_end(item["_x_right"], c_start)
        r_start = item["row_num"]
        r_end   = find_row_end(item["_y_bottom"], r_start)

        item["col_num"]     = col_num_to_letter(c_start)
        item["col_num_end"] = col_num_to_letter(c_end)
        item["row_num_end"] = r_end

    return items, col_display_widths


# ─── 输出文件一：_table.xlsx ──────────────────────────────────────────────────

def save_table_xlsx(pages_data: list[tuple], out_path: str):
    """
    按行列号将文本写入 Excel，每页一个 Sheet，自动合并跨行/列单元格。

    pages_data: list of (items, col_boundaries)
      items          : enriched list（含 row_num, row_num_end, col_num, col_num_end）
      col_boundaries : list of [x_start, x_end]
    """
    wb = Workbook()
    wb.remove(wb.active)   # 删除默认空 Sheet

    for page_idx, (items, col_boundaries) in enumerate(pages_data):
        if not items:
            continue

        ws = wb.create_sheet(title=f"第{page_idx + 1}页")

        # 设置列宽（x 跨度 / 7 ≈ Excel 字符单位，最小 8）
        for col_idx, (x_start, x_end) in enumerate(col_boundaries, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(8, (x_end - x_start) / 7)

        for item in items:
            r1 = item["row_num"]
            r2 = item["row_num_end"]
            c1 = letter_to_col_idx(item["col_num"])
            c2 = letter_to_col_idx(item["col_num_end"])

            # 先 merge，再写入锚点单元格（anchor = 左上角），跳过已被 merge 的单元格
            if r2 > r1 or c2 > c1:
                start_ref = f"{get_column_letter(c1)}{r1}"
                end_ref   = f"{get_column_letter(c2)}{r2}"
                try:
                    ws.merge_cells(f"{start_ref}:{end_ref}")
                except Exception:
                    pass   # 与已有合并区域冲突时静默跳过

            # 写入值：检查单元格是否为 MergedCell（只读副格），若是则跳过
            raw = ws.cell(row=r1, column=c1)
            try:
                raw.value = item.get("text", "")
                raw.alignment = Alignment(wrap_text=True, vertical="center")
            except AttributeError:
                pass   # MergedCell 副格，已由锚点单元格承载内容

    wb.save(out_path)
    print(f"  还原格式表格 → {out_path}")


# ─── 输出文件二：增强 JSON ────────────────────────────────────────────────────

def enrich_json(json_path: str, items: list):
    """读取已保存的 JSON，追加 row_nums / col_nums 字段后覆写。"""
    if not os.path.exists(json_path):
        print(f"  [警告] JSON 不存在，跳过增强：{json_path}")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    data["row_nums"] = [item.get("row_num", 0)    for item in items]
    data["col_nums"] = [item.get("col_num",  "A") for item in items]

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"  增强 JSON（row_nums/col_nums）→ {json_path}")


# ─── 输出文件三：_data.xlsx ───────────────────────────────────────────────────

def save_data_xlsx(all_items: list, out_path: str, threshold: float):
    """
    将所有页面的识别结果写入 Excel 数据表，低置信度行标红背景。

    列：页码 | 序号 | 识别文本 | 置信度 | 左上X | 左上Y | 右下X | 右下Y | 行号 | 列号（起） | 列号（止）| 行号（止）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR数据"

    headers = ["页码", "序号", "识别文本", "置信度",
               "左上X", "左上Y", "右下X", "右下Y",
               "行号", "列号", "列号止", "行号止"]
    ws.append(headers)

    # 表头加粗
    for cell in ws[1]:
        cell.font = Font(bold=True)

    red_fill = PatternFill(
        start_color=LOW_CONF_FILL_COLOR,
        end_color=LOW_CONF_FILL_COLOR,
        fill_type="solid",
    )

    seq = 0
    for item in all_items:
        seq += 1
        poly = item.get("poly", [])
        x0 = int(poly[0][0]) if len(poly) >= 1 else ""
        y0 = int(poly[0][1]) if len(poly) >= 1 else ""
        x2 = int(poly[2][0]) if len(poly) >= 3 else ""
        y2 = int(poly[2][1]) if len(poly) >= 3 else ""

        row_data = [
            item.get("page",  1),
            seq,
            item.get("text",  ""),
            item.get("score", 0.0),
            x0, y0, x2, y2,
            item.get("row_num",     ""),
            item.get("col_num",     ""),
            item.get("col_num_end", ""),
            item.get("row_num_end", ""),
        ]
        ws.append(row_data)

        if item.get("score", 1.0) < threshold:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    # 列宽
    col_widths = [6, 6, 30, 8, 8, 8, 8, 8, 6, 8, 8, 8]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)
    print(f"  数据存储表格 → {out_path}")


# ─── 辅助函数 ─────────────────────────────────────────────────────────────────

def _get_field(res, field_name):
    if hasattr(res, "get"):
        val = res.get(field_name, [])
    else:
        val = getattr(res, field_name, [])
    if isinstance(val, np.ndarray):
        return val.tolist()
    return list(val)


# ─── 主流程 ───────────────────────────────────────────────────────────────────

def ocr_pdf_to_excel(
    pdf_path: str,
    output_dir: str,
    threshold: float = 0.991,
    row_tol: int = 15,
):
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    print(f"正在将 PDF 转为图片：{pdf_path}")
    images = pdf_to_images(pdf_path)

    ocr = PaddleOCR(
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
    )

    pdf_name    = Path(pdf_path).stem
    all_items   = []       # 跨页汇总，用于 _data.xlsx
    pages_data  = []       # [(enriched_items, col_boundaries)] 每页，用于 _table.xlsx
    page_timings: list[dict] = []   # 每页各阶段耗时记录

    total_start = time.perf_counter()

    for page_idx, img in enumerate(images):
        page_no   = page_idx + 1
        t_page_start = time.perf_counter()

        # ── 渲染保存
        img_path = os.path.join(output_dir, f"{pdf_name}_page_{page_no}.jpg")
        img.save(img_path, quality=95)
        t_render = time.perf_counter()

        print(f"\n{'─'*50}")
        print(f"第 {page_no}/{len(images)} 页  开始时间：{time.strftime('%H:%M:%S')}")

        # ── OCR 识别
        results = ocr.predict(input=img_path)
        t_ocr = time.perf_counter()

        low_conf_items = []
        page_items     = []

        for res in results:
            res.save_to_img(output_dir)
            res.save_to_json(output_dir)

            rec_texts  = _get_field(res, "rec_texts")
            rec_scores = _get_field(res, "rec_scores")
            dt_polys   = _get_field(res, "dt_polys")

            for i, (text, score) in enumerate(zip(rec_texts, rec_scores)):
                score_f = round(float(score), 4)
                poly    = dt_polys[i] if i < len(dt_polys) else []
                item = {
                    "page":  page_no,
                    "text":  text,
                    "score": score_f,
                    "poly":  poly,
                }
                page_items.append(item)
                all_items.append(item)

                if score_f < threshold:
                    low_conf_items.append((text, score_f, poly))

        # ── 打勾检测
        if _CV2_AVAILABLE:
            detect_checkboxes(page_items, img_path)
        t_checkbox = time.perf_counter()

        # ── 低置信度标注图（线框版 + 热力图版）
        low_conf_path = os.path.join(output_dir, f"{pdf_name}_page_{page_no}_low_conf.jpg")
        heat_path     = os.path.join(output_dir, f"{pdf_name}_page_{page_no}_conf_heat.jpg")
        if low_conf_items:
            print(f"  {len(low_conf_items)} 处低于 {threshold * 100:.1f}%，正在标注…")
            draw_low_conf_boxes(img_path, low_conf_items, low_conf_path)
            draw_conf_heatmap(img_path, low_conf_items, threshold, heat_path)
        else:
            print(f"  所有文本置信度均达标（≥ {threshold * 100:.1f}%）")
        t_annotation = time.perf_counter()

        # ── 行列聚类 + 增强 JSON
        if page_items:
            enriched, col_boundaries = assign_row_col(page_items, row_tol)
            pages_data.append((enriched, col_boundaries))
            json_path = os.path.join(output_dir, f"{pdf_name}_page_{page_no}_res.json")
            enrich_json(json_path, enriched)
        else:
            pages_data.append(([], []))
        t_cluster = time.perf_counter()

        # ── 本页耗时汇总
        elapsed_total  = t_cluster    - t_page_start
        elapsed_render = t_render     - t_page_start
        elapsed_ocr    = t_ocr        - t_render
        elapsed_cb     = t_checkbox   - t_ocr
        elapsed_annot  = t_annotation - t_checkbox
        elapsed_json   = t_cluster    - t_annotation

        page_timings.append({
            "page":       page_no,
            "total_s":    elapsed_total,
            "render_s":   elapsed_render,
            "ocr_s":      elapsed_ocr,
            "checkbox_s": elapsed_cb,
            "annot_s":    elapsed_annot,
            "cluster_s":  elapsed_json,
            "texts":      len(page_items),
        })

        print(
            f"  耗时：渲染 {elapsed_render:.1f}s | "
            f"OCR {elapsed_ocr:.1f}s | "
            f"打勾检测 {elapsed_cb:.1f}s | "
            f"标注图 {elapsed_annot:.1f}s | "
            f"聚类/JSON {elapsed_json:.1f}s | "
            f"本页合计 {elapsed_total:.1f}s"
        )
        print(f"  识别文本：{len(page_items)} 条  "
              f"低置信度：{len(low_conf_items)} 条")

    # ── 三文件输出
    t_xlsx_start = time.perf_counter()
    if all_items:
        print(f"\n{'─'*50}")
        print(f"共识别 {len(all_items)} 条文本，正在生成输出文件…")

        table_path = os.path.join(output_dir, f"{pdf_name}_table.xlsx")
        save_table_xlsx(pages_data, table_path)

        data_path = os.path.join(output_dir, f"{pdf_name}_data.xlsx")
        save_data_xlsx(all_items, data_path, threshold)

        t_xlsx_end   = time.perf_counter()
        total_elapsed = t_xlsx_end - total_start

        # ── 最终汇总
        print(f"\n{'═'*50}")
        print(f"✓ 全部完成！输出目录：{output_dir}")
        print(f"{'─'*50}")
        print(f"{'页码':^6} {'文本数':^7} {'OCR':^8} {'本页合计':^10}")
        print(f"{'─'*50}")
        for pt in page_timings:
            print(f"  第{pt['page']:>2}页   {pt['texts']:>5}条   "
                  f"{pt['ocr_s']:>6.1f}s   {pt['total_s']:>7.1f}s")
        if len(page_timings) > 1:
            print(f"{'─'*50}")
            total_ocr = sum(p["ocr_s"]   for p in page_timings)
            total_pg  = sum(p["total_s"] for p in page_timings)
            print(f"  合计     {len(all_items):>5}条   "
                  f"{total_ocr:>6.1f}s   {total_pg:>7.1f}s")
        print(f"{'─'*50}")
        print(f"  xlsx 输出：{t_xlsx_end - t_xlsx_start:.1f}s")
        print(f"  总耗时：{total_elapsed:.1f}s  "
              f"（{time.strftime('%H:%M:%S')} 结束）")
        print(f"{'─'*50}")
        print(f"  {pdf_name}_table.xlsx           — 还原格式表格（含单元格合并）")
        print(f"  {pdf_name}_data.xlsx            — 数据存储表格")
        print(f"  {pdf_name}_page_N_res.json      — 增强 JSON（每页一份）")
        print(f"  {pdf_name}_page_N_low_conf.jpg  — 低置信度线框标注图（每页）")
        print(f"  {pdf_name}_page_N_conf_heat.jpg — 置信度热力图（每页）")
        print(f"{'═'*50}")
    else:
        print("\n未识别到任何内容")


if __name__ == "__main__":
    args = parse_args()
    ocr_pdf_to_excel(args.input, args.output, args.threshold, args.row_tol)
