"""
机械性能提取 — 主流程编排

将各步骤（渲染 → OCR → 筛选 → Excel → 判定 → 标注）串联为完整流程。
外部模块可直接调用 extract_mechanical() 即可。
"""

import shutil
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_writer import html_table_to_sheet
from .field_extractor import extract_named_fields
from .image_marker import mark_failures_on_image
from .judgment import apply_value_judgment
from .ocr_engine import create_engine, extract_tables_from_page
from .preprocess import image_to_pages, pdf_to_images
from .section_filter import (
    extract_mechanical_section_html,
    html_table_text,
    is_mechanical_table,
)

try:
    from ocr_config import load_config
except ImportError:
    def load_config():
        return {"output": {"output_dir": "output2"}}

try:
    from ocr_db import build_record, ensure_nonconformity_table, insert_nonconformity_records
except ImportError:
    def ensure_nonconformity_table(): pass
    def insert_nonconformity_records(rows): pass
    def build_record(**kw): return kw


def extract_mechanical(
    pdf_path: str,
    output_dir: str,
    remove_stamp: bool = True,
) -> dict:
    """
    从综合检测报告中识别并提取机械性能/力学试验表格。

    参数:
        pdf_path:    输入 PDF / 图片路径
        output_dir:  输出根目录
        remove_stamp: 是否自动去除红色公章

    返回:
        {
            "xlsx_path": str | None,
            "fields":    list[dict],   # 结构化提取字段
            "failures":  list[dict],   # 不合格项
            "marked":    int,          # 标注图数量
        }
    """
    from PIL import Image

    pdf_name = Path(pdf_path).stem
    total_start = time.perf_counter()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    report_type = "机械性能"

    out_dir = Path(output_dir) / f"{pdf_name}_机械性能_{timestamp}"
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'═'*60}")
    print(f"  [机械性能专项提取]  {Path(pdf_path).name}")
    print(f"  输出目录：{out_dir}")
    print(f"  时间戳：{timestamp}")
    print(f"{'═'*60}")

    # ── 1. 准备图片 ─────────────────────────────────────────────────
    tmp_dir = out_dir / "_tmp"
    suffix = Path(pdf_path).suffix.lower()
    if suffix == ".pdf":
        print(f"\n[1/4] 渲染 PDF 页面{'（自动去章）' if remove_stamp else ''}…")
        pages = pdf_to_images(pdf_path, tmp_dir=tmp_dir, remove_stamp=remove_stamp)
    elif suffix in (".jpg", ".jpeg", ".png"):
        print(f"\n[1/4] 读取图片文件{'（自动去章）' if remove_stamp else ''}…")
        pages = image_to_pages(pdf_path, tmp_dir=tmp_dir, remove_stamp=remove_stamp)
    else:
        print(f"  [错误] 不支持的文件格式：{suffix}")
        return {"xlsx_path": None, "fields": [], "failures": [], "marked": 0}

    archived_pages: list[tuple[int, str]] = []
    for page_no, pth in pages:
        dest = out_dir / f"识别底图_{page_no}_{timestamp}.png"
        shutil.copy2(pth, dest)
        archived_pages.append((page_no, str(dest)))
    pages = archived_pages
    print(f"  已保存识别底图（共 {len(pages)} 张）")

    page_sizes: dict[int, tuple[int, int]] = {}
    for page_no, pth in pages:
        with Image.open(pth) as im:
            page_sizes[page_no] = im.size

    # ── 2. PPStructureV3 识别 + 筛选 ───────────────────────────────
    print("\n[2/4] 启动 PPStructureV3 并筛选机械性能表格…")
    engine = create_engine()

    all_tables: list[tuple[int, int, dict]] = []
    page_ocr_lines: dict[int, list[dict]] = {}
    total_found = 0
    mech_found = 0

    for page_no, img_path in pages:
        t0 = time.perf_counter()
        print(f"\n  第 {page_no} 页  识别中…")
        tables, ocr_lines = extract_tables_from_page(engine, img_path)
        page_ocr_lines[page_no] = ocr_lines
        elapsed = time.perf_counter() - t0
        total_found += len(tables)

        mech_tables = []
        for j, tbl_info in enumerate(tables, 1):
            html = tbl_info["html"]
            if not is_mechanical_table(html):
                print(f"    ✗ 表格 {j} — 跳过（无力学关键词）：{html_table_text(html)[:80]}…")
                continue

            section_html = extract_mechanical_section_html(html)
            if section_html is not None:
                extracted = dict(tbl_info)
                extracted["html"] = section_html
                extracted["cell_bboxes"] = None
                extracted["_full_html"] = html
                mech_tables.append((j, extracted))
                mech_found += 1
                print(f"    ✓ 表格 {j} — 从综合表中提取力学区段")
            else:
                mech_tables.append((j, tbl_info))
                mech_found += 1
                print(f"    ✓ 表格 {j} — 整表识别为机械性能表格")

        for j, tbl_info in mech_tables:
            all_tables.append((page_no, j, tbl_info))
            html_path = out_dir / f"机械性能_page{page_no}_table{j}.html"
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(f"<html><meta charset='utf-8'><body>{tbl_info['html']}</body></html>")
            if "_full_html" in tbl_info:
                full_html_path = out_dir / f"完整表格_page{page_no}_table{j}.html"
                with open(full_html_path, "w", encoding="utf-8") as f:
                    f.write(f"<html><meta charset='utf-8'><body>{tbl_info['_full_html']}</body></html>")

        print(f"  第 {page_no} 页  共 {len(tables)} 个表格，机械性能：{len(mech_tables)} 个  耗时 {elapsed:.1f}s")

    print(f"\n  汇总：全部表格 {total_found} 个，机械性能表格 {mech_found} 个")

    if not all_tables:
        print("  未检测到机械性能表格，退出。")
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir)
        return {"xlsx_path": None, "fields": [], "failures": [], "marked": 0}

    # ── 3. 提取字段 + 判定 → 简洁 Excel ─────────────────────────
    print(f"\n[3/4] 提取字段并写入 Excel…  共 {len(all_tables)} 个机械性能表格")

    page_failures_map: dict[int, list] = defaultdict(list)
    td_coord_maps: dict[int, dict] = {}
    xlsx_path = out_dir / f"表格_{report_type}_{timestamp}.xlsx"
    db_rows: list[dict] = []
    all_fields: list[dict] = []
    all_failures: list[dict] = []

    ensure_nonconformity_table()

    for idx, (page_no, tbl_no, tbl_info) in enumerate(all_tables):
        tag = f"P{page_no}_T{tbl_no}"

        tmp_wb = Workbook()
        tmp_ws = tmp_wb.active
        last_row, td_coord_map = html_table_to_sheet(tmp_ws, tbl_info["html"], row_offset=1)
        td_coord_maps[id(tbl_info)] = td_coord_map

        max_col = tmp_ws.max_column
        failures = apply_value_judgment(tmp_ws)
        print(f"  [{tag}]  {last_row-1} 行 × {max_col} 列")

        print(f"\n  ── OCR 识别内容（机械性能区段）──")
        for r in range(1, last_row):
            row_vals = []
            for c in range(1, max_col + 1):
                raw = tmp_ws.cell(row=r, column=c).value
                row_vals.append(str(raw).strip() if raw is not None else "")
            if any(v for v in row_vals):
                print(f"    行{r}: {' | '.join(row_vals)}")

        fields_list = extract_named_fields(tbl_info["html"])
        all_fields.extend(fields_list)
        for record in fields_list:
            direction = record.get("方向", "?")
            non_dash = {k: v for k, v in record.items() if k != "方向" and v != "—"}
            if not non_dash:
                continue
            print(f"\n  ── {direction} 提取结果 ──")
            for k, v in non_dash.items():
                print(f"    {k}：{v}")

        all_failures.extend(failures)
        for f in failures:
            page_failures_map[page_no].append(
                (tbl_info, f["excel_row"], f["excel_col"], f["actual"], f["reason"])
            )
            db_rows.append(
                build_record(
                    pdf_path=pdf_path, report_type=report_type,
                    xlsx_path=str(xlsx_path), batch_id=timestamp,
                    sheet_name=tag, page_no=page_no,
                    table_index=tbl_no, excel_row=f["excel_row"],
                    excel_col=f["excel_col"], sample_no=f["sample_no"],
                    test_item=f["test_item"], standard_text=f["standard_text"],
                    rule_type=f["rule_type"], actual_value=f["actual"],
                    fail_reason=f["reason"], col_letter=get_column_letter(f["excel_col"]),
                )
            )

    # ── 写入简洁格式 Excel（检测项 | 供应商数值）──────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "机械性能"

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(vertical="center", horizontal="center")

    ws.cell(row=1, column=1, value="检测项").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.cell(row=1, column=2, value="供应商数值").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.cell(row=1, column=2).alignment = CENTER

    out_row = 2
    seen_fields: set[str] = set()
    for record in all_fields:
        non_dash = {k: v for k, v in record.items() if k != "方向" and v != "—"}
        if not non_dash:
            continue
        for field_name, value in non_dash.items():
            if field_name in seen_fields:
                continue
            seen_fields.add(field_name)
            ws.cell(row=out_row, column=1, value=field_name).alignment = CENTER
            ws.cell(row=out_row, column=2, value=value).alignment = CENTER
            out_row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    wb.save(str(xlsx_path))
    insert_nonconformity_records(db_rows)
    print(f"  Excel 已保存：{xlsx_path.name}（{out_row - 2} 条检测项）")

    # ── 4. 生成不合格标注图片 ─────────────────────────────────────
    print(f"\n[4/4] 生成不合格标注图片…")
    marked_count = 0
    for page_no, img_path in pages:
        failures_this_page = page_failures_map.get(page_no, [])
        if not failures_this_page:
            print(f"  第 {page_no} 页  无不合格项，跳过。")
            continue
        marked_path = out_dir / f"错误标注_{page_no}_{timestamp}.jpg"
        mark_failures_on_image(
            img_path, failures_this_page, str(marked_path),
            td_coord_maps, ocr_lines=page_ocr_lines.get(page_no),
        )
        marked_count += 1

    if tmp_dir.exists():
        shutil.rmtree(tmp_dir)
        print("  [清理] 临时图片已删除。")

    total_elapsed = time.perf_counter() - total_start
    print(f"\n{'═'*60}")
    print(f"✓ 完成！总耗时：{total_elapsed:.1f}s")
    print(f"  {xlsx_path.name}  — 机械性能表格（含合规标注）")
    if marked_count:
        print(f"  错误标注图片 — {marked_count} 张")
    print(f"{'═'*60}")

    return {
        "xlsx_path": str(xlsx_path),
        "fields": all_fields,
        "failures": all_failures,
        "marked": marked_count,
    }
