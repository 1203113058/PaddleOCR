"""
机械性能提取 — 判定规则引擎

支持四种要求值形式：区间 / ≥ / ≤ / 精确等于。
"""

import re

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def _normalize_req(req: str) -> str:
    s = req.strip()
    for a, b in (("\u2212", "-"), ("\u2013", "-"), ("\u2014", "-"), ("－", "-")):
        s = s.replace(a, b)
    s = re.sub(r"^-\s+", "-", s)
    s = re.sub(r"^\+\s+", "+", s)
    s = re.sub(r"\s*(?:℃|°\s*C|°C)\s*$", "", s, flags=re.I)
    return s.strip()


def _normalize_token(p: str) -> str:
    s = p.strip()
    for a, b in (("\u2212", "-"), ("\u2013", "-"), ("\u2014", "-"), ("－", "-")):
        s = s.replace(a, b)
    s = re.sub(r"^-\s+", "-", s)
    s = re.sub(r"^\+\s+", "+", s)
    s = re.sub(r"\s*(?:℃|°\s*C|°C)\s*$", "", s, flags=re.I)
    return s.strip()


def parse_requirement(req: str) -> dict | None:
    """解析要求值字符串，返回 {"type": ..., "min"/"max"/"value": ...} 或 None。"""
    req = _normalize_req(req)
    if not req:
        return None
    m = re.match(r'^(\d+\.?\d*)\s*[-–]\s*(\d+\.?\d*)$', req)
    if m:
        return {"type": "range", "min": float(m.group(1)), "max": float(m.group(2))}
    m = re.match(r'^[≥>]=?\s*(-?\d+\.?\d*)$', req)
    if m:
        return {"type": "gte", "value": float(m.group(1))}
    m = re.match(r'^[≤<]=?\s*(-?\d+\.?\d*)$', req)
    if m:
        return {"type": "lte", "value": float(m.group(1))}
    m = re.match(r'^[-+]?\d+\.?\d*$', req)
    if m:
        return {"type": "eq", "value": float(req)}
    return None


def check_value(actual: str, rule: dict) -> bool | None:
    """
    判断实测值是否满足规则。
    True=合格 / False=不合格 / None=无法解析
    支持 / 或空格分隔的多值。
    """
    actual = actual.strip()
    if not actual:
        return None
    parts = re.split(r'[/\s]+', actual)
    values: list[float] = []
    for p in parts:
        p = _normalize_token(p)
        if not p:
            continue
        try:
            values.append(float(p))
        except ValueError:
            return None
    if not values:
        return None

    def _ok(v: float) -> bool:
        t = rule["type"]
        if t == "range":
            return rule["min"] <= v <= rule["max"]
        if t == "gte":
            return v >= rule["value"]
        if t == "lte":
            return v <= rule["value"]
        if t == "eq":
            return abs(v - rule["value"]) < 0.01
        return True

    return all(_ok(v) for v in values)


def infer_test_item_for_column(rows_by_row, req_excel_row, col_idx):
    skip_exact = {
        "要求值", "Standard", "standard", "单位", "Unit",
        "Unit Symbol", "实测值", "Actual", "试样编号",
        "Sample No.", "Sample No",
    }
    parts: list[str] = []
    r = req_excel_row - 1
    steps = 0
    while r >= 1 and steps < 12:
        vals = rows_by_row.get(r)
        if vals is None or col_idx >= len(vals):
            r -= 1; steps += 1; continue
        cell = vals[col_idx].strip()
        if not cell or cell in skip_exact:
            r -= 1; steps += 1; continue
        parts.append(cell)
        if len(parts) >= 8:
            break
        r -= 1; steps += 1
    if not parts:
        return f"列{get_column_letter(col_idx + 1)}"
    return " ".join(reversed(parts))[:768]


def apply_value_judgment(ws) -> list[dict]:
    """
    遍历 worksheet，找「要求值」行，按列判定实测值。
    返回不合格记录列表。
    """
    FILL_FAIL = PatternFill(fill_type="solid", fgColor="FFCCCC")
    FILL_UNKNOWN = PatternFill(fill_type="solid", fgColor="FFF2CC")
    FONT_FAIL = Font(bold=True, color="CC0000")

    max_row = ws.max_row
    max_col = ws.max_column

    rows_data: list[tuple[int, list[str]]] = []
    for r in range(1, max_row + 1):
        vals = []
        for c in range(1, max_col + 1):
            raw = ws.cell(row=r, column=c).value
            vals.append(str(raw).strip() if raw is not None else "")
        rows_data.append((r, vals))

    req_excel_row: int | None = None
    req_vals: list[str] = []
    for excel_row, row_vals in rows_data:
        if any("要求值" in v or "Standard" in v for v in row_vals):
            req_excel_row = excel_row
            req_vals = row_vals
            break

    if req_excel_row is None:
        print("  [判定] 未找到「要求值」行，跳过合规判定。")
        return []

    col_rules: dict[int, dict] = {}
    for col_idx, req in enumerate(req_vals):
        rule = parse_requirement(req)
        if rule:
            col_rules[col_idx] = rule

    if not col_rules:
        print("  [判定] 要求值行中未解析到有效规则，跳过。")
        return []

    rows_by_row = {r: vals for r, vals in rows_data}
    judged = 0
    failures: list[dict] = []
    for excel_row, row_vals in rows_data:
        if excel_row <= req_excel_row:
            continue
        first_val = next((v for v in row_vals if v), "")
        if not re.match(r'^\d{2,}[-–]\d{4,}', first_val):
            continue
        for col_idx, rule in col_rules.items():
            actual = row_vals[col_idx] if col_idx < len(row_vals) else ""
            if not actual:
                continue
            result = check_value(actual, rule)
            cell = ws.cell(row=excel_row, column=col_idx + 1)
            if result is False:
                cell.fill = FILL_FAIL
                cell.font = FONT_FAIL
                t = rule["type"]
                if t == "range":
                    reason = f"{actual}<{rule['min']}或>{rule['max']}"
                elif t == "gte":
                    reason = f"{actual}<{rule['value']}"
                elif t == "lte":
                    reason = f"{actual}>{rule['value']}"
                elif t == "eq":
                    exp = rule["value"]
                    exp_s = str(int(exp)) if float(exp).is_integer() else str(exp)
                    reason = f"{actual}≠{exp_s}"
                else:
                    reason = f"{actual}≠{rule['value']}"
                standard_text = req_vals[col_idx] if col_idx < len(req_vals) else ""
                test_item = infer_test_item_for_column(rows_by_row, req_excel_row, col_idx)
                failures.append({
                    "excel_row": excel_row, "excel_col": col_idx + 1,
                    "actual": actual, "reason": reason, "rule_type": t,
                    "standard_text": standard_text, "sample_no": first_val,
                    "test_item": test_item,
                })
                judged += 1
            elif result is None:
                cell.fill = FILL_UNKNOWN

    print(f"  [判定] 完成，不合格单元格：{judged} 个（红色标注）。")
    return failures
